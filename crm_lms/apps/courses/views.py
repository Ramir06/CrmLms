from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum
from django.utils import timezone
import json
from .models import Course, CourseStudent
from .forms import CourseForm, CourseStudentForm, AddTicketsForm, MarkAttendanceForm, AdjustTicketsForm
from apps.accounts.decorators import role_required
from apps.core.mixins import get_current_organization, courses_required
from apps.lessons.models import Lesson
import datetime
import json
from .services import TicketService
from .tickets import TicketBalance, TicketTransaction, TicketAttendance


@login_required
@courses_required
def course_list(request):
    from django.db.models import Count, Q as Qm
    # Фильтруем курсы по текущей организации
    current_org = get_current_organization(request.user)
    qs = Course.objects.filter(organization=current_org) if current_org else Course.objects.none()
    qs = qs.filter(is_archived=False).select_related('mentor').annotate(
        active_count=Count('course_students', filter=Qm(course_students__status='active')),
        left_count=Count('course_students', filter=Qm(course_students__status='left')),
        frozen_count=Count('course_students', filter=Qm(course_students__status='frozen')),
    )
    status_filter = request.GET.get('status', '')
    subject_filter = request.GET.get('subject', '')
    search = request.GET.get('q', '')

    if status_filter:
        qs = qs.filter(status=status_filter)
    if subject_filter:
        qs = qs.filter(subject__icontains=subject_filter)
    if search:
        qs = qs.filter(title__icontains=search)

    subjects = Course.objects.filter(is_archived=False).values_list('subject', flat=True).distinct()
    context = {
        'courses': qs,
        'subjects': subjects,
        'status_filter': status_filter,
        'subject_filter': subject_filter,
        'search': search,
        'page_title': 'Курсы',
    }
    return render(request, 'admin/courses/list.html', context)


@login_required
@role_required(['admin', 'superadmin'])
def course_archive(request):
    courses = Course.objects.filter(is_archived=True).select_related('mentor')
    return render(request, 'admin/courses/archive.html', {'courses': courses, 'page_title': 'Архив курсов'})


def _add_months(d, months):
    import calendar as _cal
    month = d.month - 1 + months
    year = d.year + month // 12
    month = month % 12 + 1
    day = min(d.day, _cal.monthrange(year, month)[1])
    return d.replace(year=year, month=month, day=day)


def _generate_lessons_for_course(course):
    """Generate Lesson objects for a course based on its schedule."""
    from apps.lessons.models import Lesson
    from datetime import timedelta, time

    if not course.start_date or not course.days_of_week:
        return 0

    # Маппинг дней недели на числа Python (0=Пн, 6=Вс)
    day_map = {
        'mon': 0, 'tue': 1, 'wed': 2, 'thu': 3,
        'fri': 4, 'sat': 5, 'sun': 6
    }
    
    weekdays = [day_map[day] for day in course.days_of_week if day in day_map]
    if not weekdays:
        return 0

    end_date = course.end_date or _add_months(course.start_date, course.duration_months)
    start_time = course.lesson_start_time or time(9, 0)
    end_time = course.lesson_end_time or time(10, 30)

    lessons_to_create = []
    current = course.start_date
    while current <= end_date:
        if current.weekday() in weekdays:
            lessons_to_create.append(Lesson(
                course=course,
                lesson_date=current,
                start_time=start_time,
                end_time=end_time,
                room=course.room,
                type='regular',
                status='scheduled',
                organization=course.organization,  # Устанавливаем организацию от курса
            ))
        current += timedelta(days=1)

    Lesson.objects.bulk_create(lessons_to_create)
    return len(lessons_to_create)


@login_required
@role_required(['admin', 'superadmin'])
def course_create(request):
    # Получаем текущую организацию
    current_org = get_current_organization(request.user)
    print(f"Current organization: {current_org}")
    print(f"POST data: {request.POST}")
    
    form = CourseForm(request.POST or None)
    
    # Передаем организацию в форму для фильтрации менторов
    if current_org:
        form._current_org = current_org
    
    if request.method == 'POST':
        if form.is_valid():
            # Устанавливаем организацию перед сохранением
            course = form.save(commit=False)
            if current_org:
                course.organization = current_org
            course.save()
            count = _generate_lessons_for_course(course)
            if count:
                messages.success(request, f'Курс «{course.title}» создан. Автоматически сгенерировано {count} занятий.')
            else:
                messages.success(request, f'Курс «{course.title}» создан.')
            return redirect('courses:detail', pk=course.pk)
        else:
            # Добавляем ошибки формы для отладки
            print("FORM ERRORS:", form.errors)
            print("FORM NON-FIELD ERRORS:", form.non_field_errors())
            for field, errors in form.errors.items():
                messages.error(request, f"Ошибка в поле {field}: {', '.join(errors)}")
            if form.non_field_errors():
                messages.error(request, f"Общие ошибки: {', '.join(form.non_field_errors())}")
    
    return render(request, 'admin/courses/form.html', {'form': form, 'page_title': 'Создать курс'})


@login_required
@role_required(['admin', 'superadmin'])
def course_detail(request, pk):
    course = get_object_or_404(Course, pk=pk)
    course_students = CourseStudent.objects.filter(
        course=course
    ).select_related('student').prefetch_related(
        'ticket_balance',
        'ticket_transactions',
        'ticket_attendances'
    ).order_by('student__full_name')
    
    # Separate active and left students
    active_students = course_students.filter(status='active')
    left_students = course_students.filter(status='left')
    
    # Get lessons for this course
    lessons = course.lessons.all().order_by('lesson_date', 'start_time')
    upcoming_lessons = lessons.filter(lesson_date__gte=timezone.localdate()).order_by('lesson_date', 'start_time')
    past_lessons = lessons.filter(lesson_date__lt=timezone.localdate()).order_by('-lesson_date', '-start_time')
    
    # Prepare ticket data for unlimited courses
    ticket_data = {}
    total_tickets = 0
    used_tickets = 0
    remaining_tickets = 0
    
    if course.is_unlimited:
        for cs in active_students:
            ticket_data[cs.pk] = TicketService.get_student_ticket_summary(cs)
            total_tickets += ticket_data[cs.pk]['total']
            used_tickets += ticket_data[cs.pk]['used']
            remaining_tickets += ticket_data[cs.pk]['remaining']
    
    # Считаем потраченные талоны за месяц для бесконечных курсов
    monthly_spent_data = {}
    if course.is_unlimited:
        from datetime import datetime
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        for cs in active_students:
            # Считаем consume транзакции за текущий месяц
            monthly_consumed = cs.ticket_transactions.filter(
                transaction_type='consume',
                created_at__month=current_month,
                created_at__year=current_year
            ).aggregate(
                total=Sum('quantity')
            )['total'] or 0
            
            monthly_spent_data[cs.pk] = int(monthly_consumed)
    
    # Choose template based on course type
    template_name = 'admin/courses/detail_unlimited.html' if course.is_unlimited else 'admin/courses/detail.html'
    
    context = {
        'course': course,
        'course_students': course_students,
        'active_students': active_students,
        'left_students': left_students,
        'lessons': lessons,
        'upcoming_lessons': upcoming_lessons,
        'past_lessons': past_lessons,
        'ticket_data': ticket_data,
        'monthly_spent_data': monthly_spent_data,
        'total_tickets': total_tickets,
        'used_tickets': used_tickets,
        'remaining_tickets': remaining_tickets,
        'page_title': course.title,
        # Добавляем данные о студентах в формате JSON для JavaScript
        'enrollments': [
            {
                'id': cs.pk,
                'student_name': cs.student.full_name,
                'student_id': cs.student.pk,
                'remaining_tickets': ticket_data.get(cs.pk, {}).get('remaining', 0),
                'total_tickets': ticket_data.get(cs.pk, {}).get('total', 0),
                'used_tickets': ticket_data.get(cs.pk, {}).get('used', 0),
            }
            for cs in active_students
        ]
    }
    return render(request, template_name, context)


@login_required
@role_required(['admin', 'superadmin'])
@require_POST
def mark_attendance_quick(request, pk):
    """Быстрая отметка посещения со страницы курса"""
    course = get_object_or_404(Course, pk=pk)
    
    try:
        data = json.loads(request.body)
        enrollment_id = data.get('enrollment_id')
        date = data.get('date')
        status = data.get('status')
        
        if not all([enrollment_id, date, status]):
            return JsonResponse({'success': False, 'error': 'Missing required data'})
        
        # Получаем студента
        enrollment = get_object_or_404(CourseStudent, id=enrollment_id, course=course)
        
        # Создаем или обновляем запись посещаемости
        from apps.attendance.models import AttendanceRecord
        from apps.lessons.models import Lesson
        
        # Ищем или создаем урок для этой даты
        lesson, created = Lesson.objects.get_or_create(
            course=course,
            lesson_date=date,
            defaults={
                'title': f'Занятие {date}',
                'start_time': datetime.time(10, 0),
                'end_time': datetime.time(11, 0),
                'type': 'regular',
                'status': 'completed',
                'created_by': request.user
            }
        )
        
        # Сохраняем посещаемость
        attendance_record, created = AttendanceRecord.objects.update_or_create(
            lesson=lesson,
            student=enrollment.student,
            defaults={'attendance_status': status, 'marked_by': request.user}
        )
        
        # Списание талонов для бесконечных курсов
        if course.is_unlimited and status == 'present' and created:
            try:
                transaction_obj, attendance = TicketService.consume_tickets(
                    enrollment=enrollment,
                    lessons_count=1,
                    marked_by=request.user,
                    lesson_date=lesson.lesson_date,
                    comment=f"Быстрая отметка посещения {lesson.id}"
                )
            except Exception as e:
                print(f"Error deducting ticket: {e}")
        
        return JsonResponse({'success': True})
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['admin', 'superadmin'])
@require_POST
def adjust_tickets_ajax(request, pk):
    """Корректировка баланса талонов студента"""
    course = get_object_or_404(Course, pk=pk)
    
    try:
        data = json.loads(request.body)
        enrollment_id = data.get('enrollment_id')
        count = data.get('count')
        action = data.get('action')  # 'add' или 'deduct'
        
        if not all([enrollment_id, count, action]):
            return JsonResponse({'success': False, 'error': 'Missing required data'})
        
        enrollment = get_object_or_404(CourseStudent, id=enrollment_id, course=course)
        
        if action == 'add':
            # Добавляем талоны
            transaction_obj = TicketService.add_tickets(
                enrollment=enrollment,
                lessons_count=count,
                marked_by=request.user,
                comment=f"Корректировка баланса: +{count}"
            )
        elif action == 'deduct':
            # Списываем талоны
            transaction_obj = TicketService.consume_tickets(
                enrollment=enrollment,
                lessons_count=count,
                marked_by=request.user,
                comment=f"Корректировка баланса: -{count}"
            )
        else:
            return JsonResponse({'success': False, 'error': 'Invalid action'})
        
        return JsonResponse({'success': True})
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@role_required(['admin', 'superadmin'])
def course_edit(request, pk):
    # Фильтруем курс по организации пользователя
    current_org = get_current_organization(request.user)
    course_queryset = Course.objects.filter(organization=current_org) if current_org else Course.objects.none()
    course = get_object_or_404(course_queryset, pk=pk)
    
    form = CourseForm(request.POST or None, instance=course)
    
    # Передаем организацию в форму для фильтрации менторов
    if current_org:
        form._current_org = current_org
    
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Курс обновлён.')
        return redirect('courses:detail', pk=pk)
    return render(request, 'admin/courses/form.html', {'form': form, 'course': course, 'page_title': 'Редактировать курс'})


@login_required
@role_required(['admin', 'superadmin', 'mentor'])
def course_add_student(request, pk):
    course = get_object_or_404(Course, pk=pk)
    
    # Проверяем права доступа: ментор может добавлять студентов только на свои курсы
    if request.user.role == 'mentor' and course.mentor != request.user:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Вы можете добавлять студентов только на свои курсы")
    
    form = CourseStudentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        cs = form.save(commit=False)
        cs.course = course
        cs.save()
        messages.success(request, 'Студент добавлен в курс.')
        return redirect('courses:detail', pk=pk)
    return render(request, 'admin/courses/add_student.html', {'form': form, 'course': course, 'page_title': 'Добавить студента'})


@login_required
@role_required(['admin', 'superadmin'])
def course_delete(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        title = course.title
        course.delete()
        messages.success(request, f'Курс «{title}» удалён.')
        return redirect('courses:list')
    return render(request, 'admin/courses/confirm_delete.html', {'course': course})


@login_required
@role_required(['admin', 'superadmin'])
def remove_student(request, pk):
    """Remove a CourseStudent enrollment."""
    from django.utils import timezone
    cs = get_object_or_404(CourseStudent, pk=pk)
    course_pk = cs.course_id
    if request.method == 'POST':
        name = cs.student.full_name
        cs.status = 'left'
        cs.left_at = timezone.now().date()
        cs.save()
        messages.success(request, f'Студент «{name}» покинул курс.')
    return redirect('courses:detail', pk=course_pk)


@login_required
def student_drawer(request, pk):
    """Partial template for student drawer in course detail."""
    cs = get_object_or_404(CourseStudent, pk=pk)
    if request.user.role not in ('admin', 'superadmin') and request.user != cs.course.mentor:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    from apps.attendance.models import AttendanceRecord
    attendance_records = AttendanceRecord.objects.filter(
        student=cs.student, lesson__course=cs.course
    ).select_related('lesson').order_by('-lesson__lesson_date')[:10]

    context = {
        'cs': cs,
        'attendance_records': attendance_records,
    }
    return render(request, 'partials/student_drawer.html', context)


# ===== TICKET VIEWS =====

@login_required
@role_required(['admin', 'superadmin'])
def add_tickets(request, cs_pk):
    """Добавить талоны студенту"""
    cs = get_object_or_404(CourseStudent, pk=cs_pk)
    
    if not cs.course.is_unlimited:
        messages.error(request, 'Талоны можно добавлять только для бесконечных курсов')
        return redirect('courses:detail', pk=cs.course.pk)
    
    if request.method == 'POST':
        form = AddTicketsForm(request.POST)
        if form.is_valid():
            try:
                transaction_obj = TicketService.add_tickets(
                    enrollment=cs,
                    quantity=form.cleaned_data['quantity'],
                    price_per_ticket=form.cleaned_data['price_per_ticket'],
                    created_by=request.user,
                    comment=form.cleaned_data['comment']
                )
                messages.success(
                    request, 
                    f'Добавлено {form.cleaned_data["quantity"]} талонов для {cs.student.full_name}'
                )
            except Exception as e:
                messages.error(request, f'Ошибка: {str(e)}')
            return redirect('courses:detail', pk=cs.course.pk)
    else:
        form = AddTicketsForm()
    
    context = {
        'form': form,
        'cs': cs,
        'page_title': f'Добавить талоны - {cs.student.full_name}'
    }
    return render(request, 'admin/courses/add_tickets.html', context)


@login_required
@role_required(['admin', 'superadmin', 'mentor'])
def mark_attendance(request, cs_pk):
    """Отметить посещение"""
    cs = get_object_or_404(CourseStudent, pk=cs_pk)
    
    # Проверяем что ментор может отмечать только на своих курсах
    if request.user.role == 'mentor' and cs.course.mentor != request.user:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Вы можете отмечать посещаемость только на своих курсах")
    
    if not cs.course.is_unlimited:
        messages.error(request, 'Посещение по талонам доступно только для бесконечных курсов')
        return redirect('courses:detail', pk=cs.course.pk)
    
    if request.method == 'POST':
        form = MarkAttendanceForm(request.POST)
        if form.is_valid():
            try:
                transaction_obj, attendance = TicketService.consume_tickets(
                    enrollment=cs,
                    lessons_count=form.cleaned_data['lessons_count'],
                    marked_by=request.user,
                    lesson_date=form.cleaned_data['lesson_date'],
                    comment=form.cleaned_data['comment']
                )
                remaining = TicketService.get_remaining_tickets(cs)
                if remaining < 0:
                    messages.warning(
                        request,
                        f'Отмечено {form.cleaned_data["lessons_count"]} занятий. '
                        f'Перерасход: {abs(remaining)} талон(ов)'
                    )
                else:
                    messages.success(
                        request,
                        f'Отмечено {form.cleaned_data["lessons_count"]} занятий. '
                        f'Осталось: {remaining} талон(ов)'
                    )
            except Exception as e:
                messages.error(request, f'Ошибка: {str(e)}')
            return redirect('courses:detail', pk=cs.course.pk)
    else:
        form = MarkAttendanceForm()
    
    context = {
        'form': form,
        'cs': cs,
        'remaining_tickets': TicketService.get_remaining_tickets(cs),
        'page_title': f'Отметить посещение - {cs.student.full_name}'
    }
    return render(request, 'admin/courses/mark_attendance.html', context)


@login_required
@role_required(['admin', 'superadmin'])
def ticket_history(request, cs_pk):
    """История операций с талонами"""
    cs = get_object_or_404(CourseStudent, pk=cs_pk)
    
    if not cs.course.is_unlimited:
        messages.error(request, 'История талонов доступна только для бесконечных курсов')
        return redirect('courses:detail', pk=cs.course.pk)
    
    transactions = cs.ticket_transactions.all().order_by('-created_at')
    attendances = cs.ticket_attendances.all().order_by('-lesson_date')
    
    # Получаем сводку
    summary = TicketService.get_student_ticket_summary(cs)
    
    context = {
        'cs': cs,
        'transactions': transactions,
        'attendances': attendances,
        'summary': summary,
        'page_title': f'История талонов - {cs.student.full_name}'
    }
    return render(request, 'admin/courses/ticket_history.html', context)


@login_required
@role_required(['admin', 'superadmin'])
def adjust_tickets(request, cs_pk):
    """Корректировать количество талонов"""
    cs = get_object_or_404(CourseStudent, pk=cs_pk)
    
    if not cs.course.is_unlimited:
        messages.error(request, 'Корректировка доступна только для бесконечных курсов')
        return redirect('courses:detail', pk=cs.course.pk)
    
    balance = TicketService.get_or_create_balance(cs)
    
    if request.method == 'POST':
        form = AdjustTicketsForm(request.POST)
        if form.is_valid():
            try:
                transaction_obj = TicketService.adjust_tickets(
                    enrollment=cs,
                    new_total=form.cleaned_data['new_total'],
                    created_by=request.user,
                    comment=form.cleaned_data['comment']
                )
                messages.success(
                    request,
                    f'Количество талонов скорректировано до {form.cleaned_data["new_total"]}'
                )
            except Exception as e:
                messages.error(request, f'Ошибка: {str(e)}')
            return redirect('courses:detail', pk=cs.course.pk)
    else:
        form = AdjustTicketsForm(initial={'new_total': balance.total_tickets})
    
    context = {
        'form': form,
        'cs': cs,
        'balance': balance,
        'page_title': f'Корректировать талоны - {cs.student.full_name}'
    }
    return render(request, 'admin/courses/adjust_tickets.html', context)


@login_required
@role_required(['admin', 'superadmin'])
def course_extend(request, pk):
    """Extend course by months or lessons."""
    print(f"course_extend called with pk={pk}")
    course = get_object_or_404(Course, pk=pk)
    print(f"Course found: {course}")
    
    if request.method == 'POST':
        try:
            extend_type = request.POST.get('extend_type')
            print(f"extend_type: {extend_type}")
            print(f"POST data: {request.POST}")
            
            if extend_type == 'months':
                months = int(request.POST.get('months', 0))
                print(f"months: {months}")
                if months < 1 or months > 12:
                    return JsonResponse({'success': False, 'error': 'Количество месяцев должно быть от 1 до 12'})
                
                # Увеличиваем длительность курса
                course.duration_months += months
                
                # Продлеваем дату окончания если есть
                if course.end_date:
                    course.end_date += timezone.timedelta(days=30 * months)
                
                # Создаем дополнительные уроки для курса
                lessons_created = 0
                # Рассчитываем количество дополнительных уроков
                # Предполагаем 4 урока в месяц в среднем
                additional_lessons = months * 4
                print(f"additional_lessons: {additional_lessons}")
                
                for i in range(additional_lessons):
                    # Находим последнюю дату урока
                    last_lesson = Lesson.objects.filter(
                        course=course,
                        status='scheduled'
                    ).order_by('lesson_date').last()
                    
                    if last_lesson:
                        # Создаем урок через неделю после последнего
                        new_date = last_lesson.lesson_date + timezone.timedelta(weeks=1)
                    else:
                        # Если уроков нет, начинаем со следующей недели
                        new_date = timezone.localdate() + timezone.timedelta(weeks=1)
                    
                    # Определяем день недели курса
                    if course.days_of_week:
                        course_days = course.days_of_week  # Уже список, не нужно split
                        day_mapping = {
                            'monday': 0, 'tuesday': 1, 'wednesday': 2,
                            'thursday': 3, 'friday': 4, 'saturday': 5, 'sunday': 6
                        }
                        
                        # Находим ближайший подходящий день недели
                        for day_offset in range(7):
                            check_date = new_date + timezone.timedelta(days=day_offset)
                            if check_date.weekday() in [day_mapping.get(day.lower(), 0) for day in course_days]:
                                new_date = check_date
                                break
                    
                    print(f"Creating lesson {i+1} on {new_date}")
                    Lesson.objects.create(
                        course=course,
                        title=f'Урок {course.title}',
                        lesson_date=new_date,
                        start_time=course.lesson_start_time or timezone.time(9, 0),
                        end_time=course.lesson_end_time or timezone.time(10, 30),
                        room=course.room,
                        status='scheduled',
                        temporary_mentor=course.mentor
                    )
                    lessons_created += 1
                
                course.save()
                print(f"Course saved. Lessons created: {lessons_created}")
                
                return JsonResponse({
                    'success': True,
                    'message': f'Курс продлен на {months} {get_month_word(months)}. Создано {lessons_created} дополнительных уроков.'
                })
                
            elif extend_type == 'lessons':
                lessons = int(request.POST.get('lessons', 0))
                print(f"lessons: {lessons}")
                if lessons < 1 or lessons > 50:
                    return JsonResponse({'success': False, 'error': 'Количество занятий должно быть от 1 до 50'})
                
                # Создаем указанное количество дополнительных уроков
                lessons_created = 0
                for i in range(lessons):
                    # Находим последнюю дату урока
                    last_lesson = Lesson.objects.filter(
                        course=course,
                        status='scheduled'
                    ).order_by('lesson_date').last()
                    
                    if last_lesson:
                        # Создаем урок через неделю после последнего
                        new_date = last_lesson.lesson_date + timezone.timedelta(weeks=1)
                    else:
                        # Если уроков нет, начинаем со следующей недели
                        new_date = timezone.localdate() + timezone.timedelta(weeks=1)
                    
                    # Определяем день недели курса
                    if course.days_of_week:
                        course_days = course.days_of_week  # Уже список, не нужно split
                        day_mapping = {
                            'monday': 0, 'tuesday': 1, 'wednesday': 2,
                            'thursday': 3, 'friday': 4, 'saturday': 5, 'sunday': 6
                        }
                        
                        # Находим ближайший подходящий день недели
                        for day_offset in range(7):
                            check_date = new_date + timezone.timedelta(days=day_offset)
                            if check_date.weekday() in [day_mapping.get(day.lower(), 0) for day in course_days]:
                                new_date = check_date
                                break
                    
                    print(f"Creating lesson {i+1} on {new_date}")
                    Lesson.objects.create(
                        course=course,
                        title=f'Урок {course.title}',
                        lesson_date=new_date,
                        start_time=course.lesson_start_time or timezone.time(9, 0),
                        end_time=course.lesson_end_time or timezone.time(10, 30),
                        room=course.room,
                        status='scheduled',
                        temporary_mentor=course.mentor
                    )
                    lessons_created += 1
                
                return JsonResponse({
                    'success': True,
                    'message': f'Создано {lessons} дополнительных уроков.'
                })
                
            else:
                return JsonResponse({'success': False, 'error': 'Неверный тип продления'})
                
        except Exception as e:
            print(f"Error in course_extend: {e}")
            print(f"Error type: {type(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)})
    
    print("Returning JSON error - method not allowed")
    return JsonResponse({'success': False, 'error': 'Метод не разрешен'})


def get_month_word(count):
    """Возвращает правильное склонение слова 'месяц'"""
    if count >= 11 and count <= 14:
        return 'месяцев'
    elif count % 10 == 1:
        return 'месяц'
    elif count % 10 >= 2 and count % 10 <= 4:
        return 'месяца'
    else:
        return 'месяцев'


@login_required
@role_required(['admin', 'superadmin'])
def student_excel_export(request, pk, enrollment_id):
    """Экспорт посещаемости студента в Excel"""
    course = get_object_or_404(Course, pk=pk)
    enrollment = get_object_or_404(CourseStudent, id=enrollment_id, course=course)
    
    # Получаем данные талонов
    ticket_data = TicketService.get_student_ticket_summary(enrollment)
    
    # Получаем посещаемость
    from apps.attendance.models import AttendanceRecord
    attendances = AttendanceRecord.objects.filter(
        student=enrollment.student,
        lesson__course=course
    ).select_related('lesson').order_by('lesson__lesson_date')
    
    # Создаем Excel файл
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Посещаемость {enrollment.student.full_name}"
    
    # Заголовки
    headers = ['Параметр', 'Значение']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Общая информация
    data = [
        ['ФИО', enrollment.student.full_name],
        ['Курс', course.title],
        ['Всего занятий', ticket_data['total']],
        ['Использовано', ticket_data['used']],
        ['Осталось', ticket_data['remaining']],
    ]
    
    for row, (key, value) in enumerate(data, 2):
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).font = Font(bold=True)
    
    # Таблица посещений
    start_row = len(data) + 3
    ws.cell(row=start_row, column=1, value='№ занятия')
    ws.cell(row=start_row, column=2, value='Дата')
    ws.cell(row=start_row, column=1, value='№ занятия').font = Font(bold=True)
    ws.cell(row=start_row, column=2, value='Дата').font = Font(bold=True)
    ws.cell(row=start_row, column=1).fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
    ws.cell(row=start_row, column=2).fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
    
    for i, attendance in enumerate(attendances, 1):
        date_str = attendance.lesson.lesson_date.strftime('%d.%m.%Y')
        ws.cell(row=start_row + i, column=1, value=i)
        ws.cell(row=start_row + i, column=2, value=date_str)
        
        # Цветовая маркировка
        if attendance.lesson.lesson_date < timezone.localdate():
            # Прошедшие занятия - желтый
            ws.cell(row=start_row + i, column=1).fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
            ws.cell(row=start_row + i, column=2).fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        else:
            # Текущие занятия - зеленый
            ws.cell(row=start_row + i, column=1).fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
            ws.cell(row=start_row + i, column=2).fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    
    # Автоматическая ширина колонок
    for col in range(1, 3):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="attendance_{enrollment.student.full_name.replace(" ", "_")}.xlsx"'}
    )
    wb.save(response)
    return response


@login_required
@role_required(['admin', 'superadmin'])
def overall_excel_export(request, pk):
    """Общий Excel отчёт по курсу"""
    course = get_object_or_404(Course, pk=pk)
    enrollments = CourseStudent.objects.filter(
        course=course,
        status='active'
    ).select_related('student')
    
    # Создаем Excel файл
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Общий отчёт {course.title}"
    
    # Создаем стили для границ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if course.is_unlimited:
        # Логика для бесконечных курсов (талоны)
        enrollments = enrollments.prefetch_related('ticket_balance', 'ticket_attendances')
        
        # Получаем все уникальные даты посещений по талонам
        all_dates = set()
        max_attendances = 0
        
        for enrollment in enrollments:
            # Получаем все посещения по талонам для этого студента
            ticket_attendances = enrollment.ticket_attendances.all().order_by('lesson_date')
            
            attendance_count = ticket_attendances.count()
            max_attendances = max(max_attendances, attendance_count)
            
            for attendance in ticket_attendances:
                all_dates.add(attendance.lesson_date)
        
        lesson_numbers = list(range(1, max_attendances + 1))
        
        # Заголовок
        ws.cell(row=1, column=1, value='Имя ребёнка')
        header_cell = ws.cell(row=1, column=1)
        header_cell.font = Font(bold=True, color='FFFFFF', size=12)
        header_cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = thin_border
        
        # Номера занятий в заголовке
        for col, lesson_num in enumerate(lesson_numbers, 2):
            ws.cell(row=1, column=col, value=f'Занятие {lesson_num}')
            header_cell = ws.cell(row=1, column=col)
            header_cell.font = Font(bold=True, color='FFFFFF', size=12)
            header_cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.border = thin_border
        
        # Данные по студентам
        for row, enrollment in enumerate(enrollments, 2):
            ws.cell(row=row, column=1, value=enrollment.student.full_name)
            student_cell = ws.cell(row=row, column=1)
            student_cell.font = Font(bold=True, size=11)
            student_cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            student_cell.border = thin_border
            student_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Получаем баланс талонов студента
            ticket_balance = getattr(enrollment, 'ticket_balance', None)
            if not ticket_balance:
                continue
                
            # Получаем все посещения по талонам
            ticket_attendances = enrollment.ticket_attendances.all().order_by('lesson_date')
            
            # Получаем все транзакции в хронологическом порядке
            all_transactions = enrollment.ticket_transactions.order_by('created_at')
            
            # Создаем точную историю баланса по датам
            balance_by_date = {}
            running_balance = ticket_balance.total_tickets  # Начинаем с общего количества
            
            # Добавляем начальный баланс на дату создания баланса
            if ticket_balance.created_at:
                balance_by_date[ticket_balance.created_at.date()] = running_balance
            
            for transaction in all_transactions:
                if transaction.transaction_type == 'add':
                    running_balance += transaction.quantity
                elif transaction.transaction_type == 'consume':
                    running_balance -= transaction.quantity
                elif transaction.transaction_type == 'adjust':
                    running_balance = transaction.quantity
                balance_by_date[transaction.created_at.date()] = running_balance
            
            # Обрабатываем каждое посещение
            for idx, attendance in enumerate(ticket_attendances, 1):
                col = idx + 1  # Колонка в Excel (начиная со 2-й)
                attendance_date = attendance.lesson_date
                
                # Находим баланс НА МОМЕНТ посещения (до этого посещения)
                balance_at_time = ticket_balance.total_tickets  # Начальный баланс
                for date, balance in sorted(balance_by_date.items()):
                    if date < attendance_date:  # Только транзакции ДО даты посещения
                        balance_at_time = balance
                
                # Проверяем, хватило ли талонов на это посещение
                # Для каждого посещения нужно 1 талон
                has_enough_tickets = balance_at_time > 0  # Просто проверяем, был ли баланс положительным
                
                # Дополнительная проверка - было ли списание на эту дату
                has_consumption = False
                for transaction in all_transactions:
                    if (transaction.transaction_type == 'consume' and 
                        transaction.created_at.date() == attendance_date):
                        has_consumption = True
                        break
                
                # Если был расход в эту дату, считаем что посещение оплачено
                if has_consumption:
                    has_enough_tickets = True
                
                # Определяем цвет - более выразительные цвета
                if has_enough_tickets:
                    # Ярко-зеленый - талонов хватило
                    cell_fill = PatternFill(
                        start_color='28A745',  # Ярко-зеленый
                        end_color='28A745', 
                        fill_type='solid'
                    )
                    font_color = 'FFFFFF'  # Белый текст для контраста
                else:
                    # Ярко-красный - талонов не хватило
                    cell_fill = PatternFill(
                        start_color='DC3545',  # Ярко-красный
                        end_color='DC3545', 
                        fill_type='solid'
                    )
                    font_color = 'FFFFFF'  # Белый текст для контраста
                
                # Показываем посещение
                cell = ws.cell(row=row, column=col, value=attendance_date.strftime('%d.%m'))
                cell.fill = cell_fill
                cell.font = Font(color=font_color, bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Автоматическая ширина колонок и общие настройки
        ws.column_dimensions['A'].width = 30
        for col in range(2, len(lesson_numbers) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Добавляем общую настройку для листа
        ws.sheet_properties.tabColor = '2C3E50'
    
    else:
        # Логика для обычных курсов (посещаемость)
        from apps.attendance.models import AttendanceRecord
        enrollments = enrollments.prefetch_related('student__attendancerecord_set')
        
        # Получаем все уникальные даты посещений
        all_dates = set()
        max_attendances = 0
        
        for enrollment in enrollments:
            # Получаем все посещения со статусом "present" для этого студента
            attendance_records = AttendanceRecord.objects.filter(
                student=enrollment.student,
                lesson__course=course,
                attendance_status='present'
            ).order_by('lesson__lesson_date')
            
            attendance_count = attendance_records.count()
            max_attendances = max(max_attendances, attendance_count)
            
            for record in attendance_records:
                all_dates.add(record.lesson.lesson_date)
        
        dates = sorted(list(all_dates))
        lesson_numbers = list(range(1, max_attendances + 1))
        
        # Заголовок
        ws.cell(row=1, column=1, value='Имя ребёнка')
        ws.cell(row=1, column=1).font = Font(bold=True, color='FFFFFF')
        ws.cell(row=1, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # Номера занятий в заголовке
        for col, lesson_num in enumerate(lesson_numbers, 2):
            ws.cell(row=1, column=col, value=str(lesson_num))
            ws.cell(row=1, column=col).font = Font(bold=True, color='FFFFFF')
            ws.cell(row=1, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # Данные по студентам
        for row, enrollment in enumerate(enrollments, 2):
            ws.cell(row=row, column=1, value=enrollment.student.full_name)
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            # Получаем все посещения со статусом "present" для этого студента
            attendance_records = AttendanceRecord.objects.filter(
                student=enrollment.student,
                lesson__course=course,
                attendance_status='present'
            ).order_by('lesson__lesson_date')
            
            # Показываем посещения
            for idx, record in enumerate(attendance_records, 1):
                col = idx + 1  # Колонка в Excel (начиная со 2-й)
                
                # Зеленый цвет для обычных курсов
                cell_fill = PatternFill(
                    start_color='E8F5E8',  # Светло-зеленый
                    end_color='E8F5E8', 
                    fill_type='solid'
                )
                
                # Показываем посещение
                ws.cell(row=row, column=col, value=record.lesson.lesson_date.strftime('%d.%m'))
                ws.cell(row=row, column=col).fill = cell_fill
        
        # Автоматическая ширина колонок
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(lesson_numbers) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 12
    
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="overall_{course.title.replace(" ", "_")}.xlsx"'}
    )
    wb.save(response)
    return response


@login_required
@role_required(['admin', 'superadmin'])
def unlimited_course_excel_export(request, pk):
    """Excel отчёт для бесконечного курса с цветовой индикацией талонов"""
    course = get_object_or_404(Course, pk=pk)
    
    if not course.is_unlimited:
        return redirect('courses:detail', pk=pk)
    
    enrollments = CourseStudent.objects.filter(
        course=course,
        status='active'
    ).select_related('student').prefetch_related(
        'ticket_attendances'
    ).order_by('student__full_name')
    
    # Создаем Excel файл
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Отчёт {course.title}"
    
    # Создаем стили для границ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Собираем посещения для каждого студента
    student_attendances = {}
    max_lessons = 0
    
    for enrollment in enrollments:
        # Ищем посещения через AttendanceRecord с цветовыми статусами
        from apps.attendance.models import AttendanceRecord
        from apps.lessons.models import Lesson
        
        attendances_records = AttendanceRecord.objects.filter(
            student=enrollment.student,
            lesson__course=course
        ).select_related('lesson').order_by('lesson__lesson_date')
        
        attendance_list = []
        
        print(f"Студент {enrollment.student.full_name}: найдено {attendances_records.count()} посещений в AttendanceRecord")
        
        lesson_num = 1
        for attendance in attendances_records:
            if attendance.lesson.lesson_date:
                # Определяем цветовой статус
                color_status = attendance.color_status
                if color_status == 'green':
                    display_color = 'green'   # Были талоны - НОРМАЛЬНО
                elif color_status == 'red':
                    display_color = 'red'     # Не было талонов - ПРОБЛЕМА!
                else:
                    # Если цвет не установлен, определяем по балансу талонов НА МОМЕНТ занятия
                    ticket_balance = getattr(enrollment, 'ticket_balance', None)
                    if ticket_balance:
                        # Получаем дату урока
                        lesson_date = attendance.lesson.lesson_date
                        
                        # Получаем всего куплено и текущий остаток
                        total_tickets = ticket_balance.total_tickets
                        current_remaining = ticket_balance.remaining_tickets
                        
                        # Считаем сколько посещений было ДО этой даты
                        attendances_before = AttendanceRecord.objects.filter(
                            student=enrollment.student,
                            lesson__course=course,
                            attendance_status='present',
                            lesson__lesson_date__lt=lesson_date
                        ).count()
                        
                        # Баланс на момент = всего куплено - посещений до этой даты
                        balance_at_moment = total_tickets - attendances_before
                        
                        # Дополнительно проверяем: если студент никогда не имел талонов, то их не было
                        if total_tickets == 0:
                            display_color = 'red'  # Никогда не было талонов - ПРОБЛЕМА!
                        elif current_remaining == 0 and attendances_before >= total_tickets:
                            display_color = 'red'  # Все талоны потрачены - ПРОБЛЕМА!
                        elif balance_at_moment > 0:
                            display_color = 'green'   # Были талоны - НОРМАЛЬНО
                        else:
                            display_color = 'red'    # Не было талонов - ПРОБЛЕМА!
                    else:
                        display_color = 'red'    # Нет информации о талонах
                
                attendance_list.append({
                    'date': attendance.lesson.lesson_date,
                    'lesson_number': lesson_num,
                    'color_status': display_color
                })
                print(f"  Посещение {lesson_num}: {attendance.lesson.lesson_date}, цвет: {display_color}")
                lesson_num += 1
        
        student_attendances[enrollment.id] = attendance_list
        max_lessons = max(max_lessons, len(attendance_list))

    print(f"DEBUG: Количество студентов: {enrollments.count()}")
    print(f"DEBUG: Максимальное количество занятий: {max_lessons}")
    
    # Если нет посещений, создаем минимум 10 колонок для будущих занятий
    if max_lessons == 0:
        max_lessons = 10
    
    # Заголовок таблицы в одной строке
    headers = ['ФИО']
    
    # Добавляем номера занятий
    for i in range(1, max_lessons + 1):
        headers.append(str(i))
    
    # Применяем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF', size=12)
        cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Данные по студентам
    for row_idx, enrollment in enumerate(enrollments, 2):
        # ФИО
        ws.cell(row=row_idx, column=1, value=enrollment.student.full_name)
        cell = ws.cell(row=row_idx, column=1)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Заполняем данные по посещениям
        attendances = student_attendances.get(enrollment.id, [])
        
        for lesson_num in range(1, max_lessons + 1):
            col_idx = lesson_num + 1  # Колонка в Excel (2, 3, 4...)
            
            # Ищем посещение для этого номера занятия
            attendance_info = None
            for attendance in attendances:
                if attendance['lesson_number'] == lesson_num:
                    attendance_info = attendance
                    break
            
            if attendance_info:
                # Было посещение - показываем дату
                cell_value = attendance_info['date'].strftime('%d.%m')
                color_status = attendance_info['color_status']
                
                # Используем сохраненный цветовой статус
                if color_status == 'green':
                    # Были талоны - ЗЕЛЕНЫЙ фон
                    fill_color = 'D4EDDA'  # Светло-зеленый
                    font_color = '155724'  # Темно-зеленый
                    print(f"  -> Зеленый фон, талоны были (из БД)")
                elif color_status == 'red':
                    # Не было талонов - КРАСНЫЙ фон
                    fill_color = 'F8D7DA'  # Светло-красный
                    font_color = '721C24'  # Темно-красный
                    print(f"  -> Красный фон, талонов не было (из БД)")
                else:
                    # Статус не определен - по умолчанию красный
                    fill_color = 'F8D7DA'  # Светло-красный
                    font_color = '721C24'  # Темно-красный
                    print(f"  -> Красный фон, статус не определен")
            else:
                # Не было посещения - пустая ячейка с белым фоном
                cell_value = ''
                fill_color = 'FFFFFF'  # Белый фон
                font_color = '000000'  # Черный текст
                print(f"  -> Белый фон, нет посещения")
            
            # Записываем значение
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Применяем стили
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            cell.font = Font(color=font_color, size=10)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 15)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Отправляем файл
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="unlimited_{course.title.replace(" ", "_")}.xlsx"'}
    )
    wb.save(response)
    return response
