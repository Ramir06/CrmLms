from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
import json
import calendar
from datetime import datetime, date

from apps.core.mixins import admin_required
from apps.core.mixins_organization import get_current_organization, filter_by_organization
from .models import FinanceTransaction, FinanceCategory, FinanceAccount


from django import forms


class TransactionForm(forms.ModelForm):
    class Meta:
        model = FinanceTransaction
        fields = ['type', 'category', 'amount', 'account', 'transfer_to_account', 'description', 'transaction_date', 'branch']
        widgets = {
            'type': forms.Select(attrs={'class': 'form-select', 'id': 'transaction_type'}),
            'category': forms.Select(attrs={'class': 'form-select'}),
            'amount': forms.NumberInput(attrs={'class': 'form-control', 'step': '0.01'}),
            'account': forms.Select(attrs={'class': 'form-select'}),
            'transfer_to_account': forms.Select(attrs={'class': 'form-select'}),
            'description': forms.Textarea(attrs={'class': 'form-control', 'rows': 2}),
            'transaction_date': forms.DateInput(attrs={'class': 'form-control', 'type': 'date'}),
            'branch': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Название филиала'}),
        }


class AccountForm(forms.ModelForm):
    class Meta:
        model = FinanceAccount
        fields = ['name', 'balance', 'description']
        widgets = {
            'name': forms.TextInput(attrs={'class': 'form-control'}),
            'balance': forms.NumberInput(attrs={'class': 'form-control', 'step': '0.01'}),
            'description': forms.Textarea(attrs={'class': 'form-control', 'rows': 3}),
        }


class CategoryForm(forms.ModelForm):
    class Meta:
        model = FinanceCategory
        fields = ['name', 'type', 'color']
        widgets = {
            'name': forms.TextInput(attrs={'class': 'form-control'}),
            'type': forms.Select(attrs={'class': 'form-select'}),
            'color': forms.TextInput(attrs={'class': 'form-control', 'type': 'color'}),
        }


@login_required
@admin_required
def finance_dashboard(request):
    """Дашборд бухгалтерии с основной статистикой"""
    current_org = get_current_organization(request.user)
    
    # Получаем текущий месяц и год
    today = timezone.now().date()
    year = request.GET.get('year', today.year)
    month = request.GET.get('month', today.month)
    
    try:
        year = int(year)
        month = int(month)
    except (ValueError, TypeError):
        year = today.year
        month = today.month
    
    # Первые и последние дни месяца
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    
    # Базовый queryset с фильтрацией по организации
    qs = FinanceTransaction.objects.select_related('category', 'account', 'created_by')
    if current_org:
        qs = qs.filter(organization=current_org)
    else:
        # Если нет организации, показываем все только для суперпользователя
        if request.user.is_superuser:
            qs = qs.all()  # Суперпользователь видит все когда нет организации
        else:
            qs = qs.none()  # Остальные не видят ничего
    
    # Транзакции за текущий месяц
    month_transactions = qs.filter(transaction_date__gte=first_day, transaction_date__lte=last_day)
    
    # Статистика за месяц
    total_income = month_transactions.filter(type='income').aggregate(s=Sum('amount'))['s'] or 0
    total_expense = month_transactions.filter(type='expense').aggregate(s=Sum('amount'))['s'] or 0
    profit = total_income - total_expense
    
    # Балансы по счетам - ТОЛЬКО текущей организации
    accounts = FinanceAccount.objects.none()
    if current_org:
        accounts = FinanceAccount.objects.filter(organization=current_org)
    
    account_balances = []
    for account in accounts:
        income = qs.filter(account=account, type='income').aggregate(s=Sum('amount'))['s'] or 0
        expense = qs.filter(account=account, type='expense').aggregate(s=Sum('amount'))['s'] or 0
        balance = income - expense
        account_balances.append({
            'account': account,
            'balance': balance,
            'income': income,
            'expense': expense
        })
    
    # Последние транзакции
    recent_transactions = qs.order_by('-transaction_date')[:10]
    
    # Данные для графика (последние 30 дней)
    days_data = []
    for i in range(30):
        day = today - timezone.timedelta(days=i)
        day_income = qs.filter(transaction_date=day, type='income').aggregate(s=Sum('amount'))['s'] or 0
        day_expense = qs.filter(transaction_date=day, type='expense').aggregate(s=Sum('amount'))['s'] or 0
        days_data.append({
            'date': day.strftime('%Y-%m-%d'),
            'income': float(day_income),
            'expense': float(day_expense)
        })
    
    context = {
        'page_title': 'Бухгалтерия',
        'total_income': total_income,
        'total_expense': total_expense,
        'profit': profit,
        'account_balances': account_balances,
        'recent_transactions': recent_transactions,
        'days_data': json.dumps(list(reversed(days_data))),
        'current_month': month,
        'current_year': year,
        'month_name': calendar.month_name[month],
        'prev_month': month - 1 if month > 1 else 12,
        'prev_year': year if month > 1 else year - 1,
        'next_month': month + 1 if month < 12 else 1,
        'next_year': year if month < 12 else year + 1,
    }
    return render(request, 'admin/finance/dashboard.html', context)


@login_required
@admin_required
def transaction_list(request):
    current_org = get_current_organization(request.user)
    
    # Базовый queryset с фильтрацией по организации
    qs = FinanceTransaction.objects.select_related('category', 'account', 'created_by').all()
    
    # Фильтруем по организации - ВСЕГДА!
    if current_org:
        qs = qs.filter(organization=current_org)
    else:
        # Если нет организации, показываем все только для суперпользователя
        if request.user.is_superuser:
            qs = qs.all()  # Суперпользователь видит все когда нет организации
        else:
            qs = qs.none()  # Остальные не видят ничего
    
    # Фильтры
    type_filter = request.GET.get('type')
    account_filter = request.GET.get('account')
    category_filter = request.GET.get('category')
    branch_filter = request.GET.get('branch')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search = request.GET.get('search')

    if type_filter:
        qs = qs.filter(type=type_filter)
    if account_filter:
        qs = qs.filter(account_id=account_filter)
    if category_filter:
        qs = qs.filter(category_id=category_filter)
    if branch_filter:
        qs = qs.filter(branch__icontains=branch_filter)
    if date_from:
        qs = qs.filter(transaction_date__gte=date_from)
    if date_to:
        qs = qs.filter(transaction_date__lte=date_to)
    if search:
        qs = qs.filter(description__icontains=search)

    # Пагинация
    paginator = Paginator(qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Статистика
    total_income = qs.filter(type='income').aggregate(s=Sum('amount'))['s'] or 0
    total_expense = qs.filter(type='expense').aggregate(s=Sum('amount'))['s'] or 0
    balance = total_income - total_expense

    # Фильтры для формы - ТОЛЬКО для текущей организации
    accounts = FinanceAccount.objects.none()
    categories = FinanceCategory.objects.none()
    
    if current_org:
        accounts = FinanceAccount.objects.filter(organization=current_org)
        categories = FinanceCategory.objects.filter(organization=current_org)

    # Получаем уникальные филиалы для фильтра (только текущей организации)
    branches = []
    if current_org:
        branches = list(FinanceTransaction.objects.filter(
            organization=current_org,
            branch__isnull=False
        ).exclude(branch='').values_list('branch', flat=True).order_by('branch').distinct())
    else:
        # Если нет организации, показываем все филиалы (только для суперпользователя)
        if request.user.is_superuser:
            branches = list(FinanceTransaction.objects.filter(
                branch__isnull=False
            ).exclude(branch='').values_list('branch', flat=True).order_by('branch').distinct())
    
    context = {
        'page_title': 'Транзакции',
        'page_obj': page_obj,
        'transactions': page_obj,  # Добавляем для совместимости с шаблоном
        'total_income': total_income,
        'total_expense': total_expense,
        'balance': balance,
        'accounts': accounts,
        'categories': categories,
        'branches': branches,
        'type_filter': type_filter,
        'account_filter': account_filter,
        'category_filter': category_filter,
        'branch_filter': branch_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search': search,
    }
    return render(request, 'admin/finance/list.html', context)


@login_required
@admin_required
def transaction_create(request):
    current_org = get_current_organization(request.user)
    
    # Получаем тип из GET параметра
    initial_data = {}
    transaction_type = request.GET.get('type')
    if transaction_type in ['income', 'expense', 'transfer']:
        initial_data['type'] = transaction_type
    
    form = TransactionForm(request.POST or None, initial=initial_data)
    
    # Фильтруем категории и счета по организации
    if current_org:
        form.fields['category'].queryset = FinanceCategory.objects.filter(organization=current_org)
        form.fields['account'].queryset = FinanceAccount.objects.filter(organization=current_org)
        form.fields['transfer_to_account'].queryset = FinanceAccount.objects.filter(organization=current_org)
    else:
        # Если нет организации, не показываем никаких вариантов
        form.fields['category'].queryset = FinanceCategory.objects.none()
        form.fields['account'].queryset = FinanceAccount.objects.none()
        form.fields['transfer_to_account'].queryset = FinanceAccount.objects.none()
    
    if request.method == 'POST' and form.is_valid():
        tx = form.save(commit=False)
        tx.created_by = request.user
        
        # Всегда устанавливаем организацию
        if current_org:
            tx.organization = current_org
        else:
            # Если нет организации, не сохраняем транзакцию
            messages.error(request, 'Не удалось определить организацию.')
            return render(request, 'admin/finance/transaction_form.html', {
                'form': form, 
                'page_title': 'Добавить транзакцию',
                'action': 'create'
            })
        
        tx.save()
        messages.success(request, 'Транзакция добавлена.')
        return redirect('finance:list')
    
    context = {
        'form': form, 
        'page_title': 'Добавить транзакцию',
        'action': 'create'
    }
    return render(request, 'admin/finance/transaction_form.html', context)


@login_required
@admin_required
def transaction_edit(request, pk):
    current_org = get_current_organization(request.user)
    transaction = get_object_or_404(FinanceTransaction, pk=pk)
    
    # Проверка организации
    if current_org and transaction.organization != current_org:
        messages.error(request, 'У вас нет доступа к этой транзакции.')
        return redirect('finance:list')
    
    # Проверка на редактирование авто-сгенерированных транзакций
    if transaction.auto_generated:
        messages.error(request, 'Нельзя редактировать автоматически созданные транзакции.')
        return redirect('finance:list')
    
    form = TransactionForm(request.POST or None, instance=transaction)
    
    # Фильтруем категории и счета по организации
    if current_org:
        form.fields['category'].queryset = FinanceCategory.objects.filter(organization=current_org)
        form.fields['account'].queryset = FinanceAccount.objects.filter(organization=current_org)
        form.fields['transfer_to_account'].queryset = FinanceAccount.objects.filter(organization=current_org)
    
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Транзакция обновлена.')
        return redirect('finance:list')
    
    context = {
        'form': form, 
        'page_title': 'Редактировать транзакцию',
        'action': 'edit',
        'transaction': transaction
    }
    return render(request, 'admin/finance/transaction_form.html', context)


@login_required
@admin_required
@require_POST
def transaction_delete(request, pk):
    current_org = get_current_organization(request.user)
    transaction = get_object_or_404(FinanceTransaction, pk=pk)
    
    # Проверка организации
    if current_org and transaction.organization != current_org:
        return JsonResponse({'error': 'У вас нет доступа к этой транзакции.'}, status=403)
    
    # Проверка на удаление авто-сгенерированных транзакций
    if transaction.auto_generated:
        return JsonResponse({'error': 'Нельзя удалить автоматически созданные транзакции.'}, status=400)
    
    transaction.delete()
    return JsonResponse({'success': True})


@login_required
@admin_required
def account_list(request):
    current_org = get_current_organization(request.user)
    
    # Строгая фильтрация по организации
    accounts = FinanceAccount.objects.none()
    if current_org:
        accounts = FinanceAccount.objects.filter(organization=current_org)
    
    # Подсчитываем баланс для каждого счета
    for account in accounts:
        transactions = FinanceTransaction.objects.filter(account=account, organization=current_org)
        
        income = transactions.filter(type='income').aggregate(s=Sum('amount'))['s'] or 0
        expense = transactions.filter(type='expense').aggregate(s=Sum('amount'))['s'] or 0
        account.calculated_balance = income - expense
        account.transaction_count = transactions.count()
    
    context = {
        'page_title': 'Счета',
        'accounts': accounts
    }
    return render(request, 'admin/finance/accounts.html', context)


@login_required
@admin_required
def account_create(request):
    current_org = get_current_organization(request.user)
    form = AccountForm(request.POST or None)
    
    if request.method == 'POST' and form.is_valid():
        account = form.save(commit=False)
        if current_org:
            account.organization = current_org
        account.save()
        messages.success(request, 'Счёт создан.')
        return redirect('finance:accounts')
    
    context = {
        'form': form,
        'page_title': 'Добавить счёт'
    }
    return render(request, 'admin/finance/account_form.html', context)


@login_required
@admin_required
def category_list(request):
    current_org = get_current_organization(request.user)
    
    # Строгая фильтрация по организации
    categories = FinanceCategory.objects.none()
    if current_org:
        categories = FinanceCategory.objects.filter(organization=current_org)
    
    # Подсчитываем количество транзакций для каждой категории
    for category in categories:
        transactions = FinanceTransaction.objects.filter(category=category, organization=current_org)
        category.transaction_count = transactions.count()
        category.total_amount = transactions.aggregate(s=Sum('amount'))['s'] or 0
    
    context = {
        'page_title': 'Категории',
        'categories': categories
    }
    return render(request, 'admin/finance/categories.html', context)


@login_required
@admin_required
def category_create(request):
    current_org = get_current_organization(request.user)
    form = CategoryForm(request.POST or None)
    
    if request.method == 'POST' and form.is_valid():
        category = form.save(commit=False)
        if current_org:
            category.organization = current_org
        category.save()
        messages.success(request, 'Категория создана.')
        return redirect('finance:categories')
    
    context = {
        'form': form,
        'page_title': 'Добавить категорию'
    }
    return render(request, 'admin/finance/category_form.html', context)


@login_required
@admin_required
def finance_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Фильтруем по организации
    current_org = get_current_organization(request.user)
    transactions = FinanceTransaction.objects.select_related('category', 'account', 'created_by').all()
    if current_org:
        transactions = transactions.filter(organization=current_org)
    
    # Применяем фильтры
    type_filter = request.GET.get('type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if type_filter:
        transactions = transactions.filter(type=type_filter)
    if date_from:
        transactions = transactions.filter(transaction_date__gte=date_from)
    if date_to:
        transactions = transactions.filter(transaction_date__lte=date_to)
    
    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Бухгалтерия'
    
    # Заголовки
    headers = ['ID', 'Дата', 'Тип', 'Категория', 'Сумма', 'Счёт', 'Счёт получения', 'Описание', 'Автор', 'Авто']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Данные
    for row, t in enumerate(transactions, 2):
        ws.cell(row=row, column=1, value=t.pk)
        ws.cell(row=row, column=2, value=str(t.transaction_date))
        ws.cell(row=row, column=3, value=t.get_type_display())
        ws.cell(row=row, column=4, value=t.category.name if t.category else '')
        ws.cell(row=row, column=5, value=float(t.amount))
        ws.cell(row=row, column=6, value=t.account.name if t.account else '')
        ws.cell(row=row, column=7, value=t.transfer_to_account.name if t.transfer_to_account else '')
        ws.cell(row=row, column=8, value=t.description)
        ws.cell(row=row, column=9, value=t.created_by.get_full_name() if t.created_by else '')
        ws.cell(row=row, column=10, value='Да' if t.auto_generated else 'Нет')
    
    # Автоматическая ширина колонок
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Статистика внизу
    total_income = transactions.filter(type='income').aggregate(s=Sum('amount'))['s'] or 0
    total_expense = transactions.filter(type='expense').aggregate(s=Sum('amount'))['s'] or 0
    
    row = len(transactions) + 3
    ws.cell(row=row, column=5, value='Итого доход:').font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(total_income)).font = Font(bold=True)
    
    row += 1
    ws.cell(row=row, column=5, value='Итого расход:').font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(total_expense)).font = Font(bold=True)
    
    row += 1
    ws.cell(row=row, column=5, value='Прибыль:').font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(total_income - total_expense)).font = Font(bold=True)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'finance_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@login_required
@admin_required
def finance_stats_api(request):
    """API для статистики в реальном времени"""
    current_org = get_current_organization(request.user)
    
    # Базовый queryset
    qs = FinanceTransaction.objects.all()
    if current_org:
        qs = qs.filter(organization=current_org)
    
    # Период
    period = request.GET.get('period', 'month')
    today = timezone.now().date()
    
    if period == 'week':
        start_date = today - timezone.timedelta(days=7)
    elif period == 'quarter':
        start_date = today - timezone.timedelta(days=90)
    elif period == 'year':
        start_date = today - timezone.timedelta(days=365)
    else:  # month
        start_date = today.replace(day=1)
    
    period_qs = qs.filter(transaction_date__gte=start_date)
    
    # Статистика
    income = period_qs.filter(type='income').aggregate(s=Sum('amount'))['s'] or 0
    expense = period_qs.filter(type='expense').aggregate(s=Sum('amount'))['s'] or 0
    
    # По категориям
    category_stats = {}
    for cat in FinanceCategory.objects.all():
        if current_org:
            cat_qs = period_qs.filter(category=cat, organization=current_org)
        else:
            cat_qs = period_qs.filter(category=cat)
        
        if cat_qs.exists():
            category_stats[cat.name] = {
                'amount': float(cat_qs.aggregate(s=Sum('amount'))['s'] or 0),
                'count': cat_qs.count(),
                'type': cat.type
            }
    
    # Ежедневная статистика за последние 30 дней
    daily_stats = []
    for i in range(30):
        day = today - timezone.timedelta(days=i)
        day_income = qs.filter(transaction_date=day, type='income').aggregate(s=Sum('amount'))['s'] or 0
        day_expense = qs.filter(transaction_date=day, type='expense').aggregate(s=Sum('amount'))['s'] or 0
        daily_stats.append({
            'date': day.strftime('%Y-%m-%d'),
            'income': float(day_income),
            'expense': float(day_expense)
        })
    
    data = {
        'income': float(income),
        'expense': float(expense),
        'profit': float(income - expense),
        'category_stats': category_stats,
        'daily_stats': list(reversed(daily_stats)),
        'period': period,
        'start_date': str(start_date),
        'end_date': str(today)
    }
    
    return JsonResponse(data)
