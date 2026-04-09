from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.template.loader import render_to_string

from apps.core.mixins import admin_required, get_menu_position, get_student_form_fields, get_current_organization, students_required
from .models import Student, StudentNote
from .forms import StudentForm


@login_required
@students_required
def student_list(request):
    from apps.courses.models import Course
    # Фильтруем студентов по текущей организации пользователя
    current_org = get_current_organization(request.user)
    
    # Отладочная информация
    print(f"DEBUG: current_org = {current_org}")
    print(f"DEBUG: request.current_organization = {getattr(request, 'current_organization', 'None')}")
    
    qs = Student.objects.filter(organization=current_org) if current_org else Student.objects.none()
    qs = qs.select_related('user_account').prefetch_related('course_enrollments__course')
    search = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    subject_filter = request.GET.get('subject', '')

    if search:
        qs = qs.filter(Q(full_name__icontains=search) | Q(phone__icontains=search))
    if status_filter:
        qs = qs.filter(status=status_filter)
    if subject_filter:
        qs = qs.filter(course_enrollments__course__subject__icontains=subject_filter).distinct()

    subjects = Course.objects.filter(organization=current_org) if current_org else Course.objects.none()
    subjects = subjects.exclude(subject='').values_list('subject', flat=True).distinct().order_by('subject')
    
    # Separate students who left courses
    active_students = []
    left_students = []
    
    for student in qs:
        has_left_course = student.course_enrollments.filter(status='left').exists()
        if has_left_course:
            left_students.append(student)
        else:
            active_students.append(student)

    context = {
        'active_students': active_students,
        'left_students': left_students,
        'search': search,
        'status_filter': status_filter,
        'subject_filter': subject_filter,
        'subjects': subjects,
        'page_title': 'Студенты',
        'menu_position': get_menu_position(),
        'student_form_fields': get_student_form_fields(),
        'current_organization': get_current_organization(request.user),
    }
    return render(request, 'admin/students/list.html', context)


@login_required
@admin_required
def student_create(request):
    form = StudentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        # Устанавливаем организацию перед сохранением
        student = form.save(commit=False)
        current_org = get_current_organization(request.user)
        if current_org:
            student.organization = current_org
            
            # Добавляем студента как участника организации
            from apps.organizations.models import OrganizationMember, UserCurrentOrganization
            if student.user_account:  # Проверяем, есть ли у студента аккаунт пользователя
                member, created = OrganizationMember.objects.get_or_create(
                    user=student.user_account,
                    organization=current_org,
                    defaults={'role': 'member'}
                )
                if created:
                    print(f"✅ Студент {student.full_name} добавлен в организацию {current_org.name}")
                
                # Устанавливаем текущую организацию для студента
                user_current_org, created = UserCurrentOrganization.objects.get_or_create(
                    user=student.user_account,
                    defaults={'organization': current_org}
                )
                if not created:
                    user_current_org.organization = current_org
                    user_current_org.save()
                
        student.save()
        messages.success(request, f'Студент «{student.full_name}» создан.')
        return redirect('admin_students:list')
    return render(request, 'admin/students/form.html', {
        'form': form, 
        'page_title': 'Добавить студента',
        'menu_position': get_menu_position(),
        'student_form_fields': get_student_form_fields(),
        'current_organization': get_current_organization(request.user),
    })


@login_required
@admin_required
def student_detail(request, pk):
    # Фильтруем студента по организации пользователя
    current_org = get_current_organization(request.user)
    student_queryset = Student.objects.filter(organization=current_org) if current_org else Student.objects.none()
    student = get_object_or_404(student_queryset, pk=pk)
    enrollments = student.course_enrollments.select_related('course').all()
    payments = student.payments.select_related('course').order_by('-paid_at')[:10]
    debts = student.debts.filter(status='active')
    
    # Загружаем заметки
    admin_notes = student.admin_notes.select_related('admin').all()
    mentor_notes = student.notes.select_related('mentor').all()
    
    context = {
        'student': student,
        'enrollments': enrollments,
        'payments': payments,
        'debts': debts,
        'admin_notes': admin_notes,
        'mentor_notes': mentor_notes,
        'page_title': student.full_name,
        'current_organization': get_current_organization(request.user),
    }
    return render(request, 'admin/students/detail.html', context)


@login_required
@admin_required
def student_edit(request, pk):
    # Фильтруем студента по организации пользователя
    current_org = get_current_organization(request.user)
    student_queryset = Student.objects.filter(organization=current_org) if current_org else Student.objects.none()
    student = get_object_or_404(student_queryset, pk=pk)
    form = StudentForm(request.POST or None, instance=student)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Данные студента обновлены.')
        return redirect('admin_students:detail', pk=pk)
    return render(request, 'admin/students/form.html', {
        'form': form, 'student': student, 'page_title': 'Редактировать студента',
        'menu_position': get_menu_position(),
        'student_form_fields': get_student_form_fields(),
        'current_organization': get_current_organization(request.user),
    })


@login_required
@admin_required
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        name = student.full_name
        student.delete()
        messages.success(request, f'Студент «{name}» удалён.')
        return redirect('admin_students:list')
    return render(request, 'admin/students/confirm_delete.html', {'student': student})


@login_required
@admin_required
def student_block(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST' and student.user_account:
        user = student.user_account
        user.is_active = not user.is_active
        user.save()
        if user.is_active:
            messages.success(request, f'Аккаунт «{student.full_name}» разблокирован.')
        else:
            messages.warning(request, f'Аккаунт «{student.full_name}» заблокирован.')
    return redirect('admin_students:detail', pk=pk)


@login_required
@admin_required
def student_reset_password(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST' and student.user_account:
        new_password = request.POST.get('new_password', '').strip()
        if new_password:
            student.user_account.set_password(new_password)
            student.user_account.save()
            messages.success(request, f'Пароль для «{student.full_name}» сброшен.')
        else:
            messages.error(request, 'Введите новый пароль.')
    return redirect('admin_students:detail', pk=pk)


@login_required
@admin_required
def student_export(request):
    import openpyxl
    from django.utils import timezone

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Студенты'
    headers = ['ID', 'Имя', 'Телефон', 'Родитель', 'Телефон родителя', 'Источник', 'Статус', 'Дата добавления']
    ws.append(headers)

    for s in Student.objects.all():
        ws.append([
            s.pk, s.full_name, s.phone, s.parent_name, s.parent_phone,
            s.source, s.status, s.created_at.strftime('%Y-%m-%d') if s.created_at else '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=students_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


@login_required
@admin_required
def student_drawer_info(request, pk):
    """Возвращает информацию о студенте для боковой панели"""
    try:
        student = get_object_or_404(Student, pk=pk)
        
        # Оптимизированные запросы с агрегацией
        enrollments = student.course_enrollments.select_related('course').prefetch_related('course__mentor').all()
        payments = student.payments.select_related('course').order_by('-paid_at')[:10]
        debts = student.debts.filter(status='active').select_related('course')
        
        # Агрегация финансов
        from django.db.models import Sum, Count
        total_paid = student.payments.aggregate(total=Sum('amount'))['total'] or 0
        total_debt = debts.aggregate(total=Sum('total_amount'))['total'] or 0
        
        context = {
            'student': student,
            'enrollments': enrollments,
            'payments': payments,
            'debts': debts,
            'total_paid': total_paid,
            'total_debt': total_debt,
            'payments_count': student.payments.count(),
            'enrollments_count': enrollments.count(),
        }
        
        html = render_to_string('admin/students/student_drawer.html', context, request=request)
        return JsonResponse({'html': html})
    
    except Exception as e:
        # Логируем ошибку и возвращаем сообщение
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in student_drawer_info: {str(e)}")
        
        return JsonResponse({
            'html': f'''
                <div class="alert alert-danger">
                    <i class="bi bi-exclamation-triangle me-2"></i>
                    Ошибка загрузки данных студента: {str(e)}
                </div>
            '''
        })


@login_required
@admin_required
def student_detailed_excel_report(request, pk):
    """Генерация детального Excel-отчёта о студенте"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.utils import timezone
    from django.db.models import Sum, Count, Avg, Max, Min
    from apps.attendance.models import AttendanceRecord
    from apps.assignments.models import AssignmentSubmission
    from apps.payments.models import Payment
    from apps.debts.models import Debt
    import logging
    
    # Фильтруем студента по организации пользователя
    current_org = get_current_organization(request.user)
    student_queryset = Student.objects.filter(organization=current_org) if current_org else Student.objects.none()
    student = get_object_or_404(student_queryset, pk=pk)
    
    # Создаём Excel файл
    wb = openpyxl.Workbook()
    
    # Удаляем стандартный лист и создаём свои
    wb.remove(wb.active)
    
    # Стили
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')
    
    # 1. Лист с основной информацией
    ws_main = wb.create_sheet('Основная информация', 0)
    
    # Заголовок
    ws_main.merge_cells('A1:F1')
    ws_main['A1'] = f'Детальный отчёт по студенту: {student.full_name}'
    ws_main['A1'].font = Font(bold=True, size=14)
    ws_main['A1'].alignment = alignment
    
    # Основная информация
    main_data = [
        ['Параметр', 'Значение'],
        ['ФИО', student.full_name or ''],
        ['Телефон', student.phone or ''],
        ['Email', student.user_account.email if student.user_account else ''],
        ['Дата рождения', student.birth_date.strftime('%d.%m.%Y') if student.birth_date else ''],
        ['Пол', student.get_gender_display() if hasattr(student, 'get_gender_display') else ''],
        ['Источник', student.get_source_display() if student.source else ''],
        ['Статус', student.get_status_display()],
        ['Дата добавления', student.created_at.strftime('%d.%m.%Y %H:%M') if student.created_at else ''],
        ['Последний вход', student.user_account.last_login.strftime('%d.%m.%Y %H:%M') if student.user_account and student.user_account.last_login else ''],
        ['Дата первого входа', student.user_account.date_joined.strftime('%d.%m.%Y %H:%M') if student.user_account else ''],
        ['Родитель', student.parent_name or ''],
        ['Телефон родителя', student.parent_phone or ''],
        ['Примечания', student.note or ''],
    ]
    
    for row_idx, row_data in enumerate(main_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_main.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 2:  # Заголовок
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = alignment
            else:
                cell.border = border
    
    # Автоширина колонок
    for column in ws_main.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_main.column_dimensions[column_letter].width = adjusted_width
    
    # 2. Лист с курсами и оценками
    ws_courses = wb.create_sheet('Курсы и оценки', 1)
    
    # Заголовки
    course_headers = ['Курс', 'Статус', 'Дата зачисления', 'Прогресс %', 'Ментор', 'Средний балл', 'Кол-во заданий', 'Сдано заданий']
    for col_idx, header in enumerate(course_headers, 1):
        cell = ws_courses.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment
    
    # Данные по курсам
    enrollments = student.course_enrollments.select_related('course', 'course__mentor').all()
    for row_idx, enrollment in enumerate(enrollments, 2):
        # Получаем оценки и статистику по заданиям
        assignments_count = enrollment.course.assignments.count() if hasattr(enrollment.course, 'assignments') else 0
        submitted_count = AssignmentSubmission.objects.filter(
            student=student, 
            assignment__course=enrollment.course
        ).count()
        
        # Средний балл
        avg_score = AssignmentSubmission.objects.filter(
            student=student, 
            assignment__course=enrollment.course,
            score__isnull=False
        ).aggregate(avg=Avg('score'))['avg'] or 0
        
        course_data = [
            enrollment.course.title or '',
            enrollment.get_status_display(),
            enrollment.joined_at.strftime('%d.%m.%Y') if enrollment.joined_at else '',
            f'{enrollment.progress_percent or 0}%',
            f'{enrollment.course.mentor.first_name} {enrollment.course.mentor.last_name}' if enrollment.course.mentor else '',
            f'{avg_score:.1f}' if avg_score else '0',
            assignments_count,
            submitted_count,
        ]
        
        for col_idx, value in enumerate(course_data, 1):
            cell = ws_courses.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
    
    # Автоширина
    for column in ws_courses.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_courses.column_dimensions[column_letter].width = adjusted_width
    
    # 3. Лист с посещаемостью
    ws_attendance = wb.create_sheet('Посещаемость', 2)
    
    # Заголовки
    attendance_headers = ['Дата', 'Курс', 'Статус', 'Тип занятия', 'Примечание']
    for col_idx, header in enumerate(attendance_headers, 1):
        cell = ws_attendance.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment
    
    # Данные по посещаемости
    attendance_records = AttendanceRecord.objects.filter(
        student=student
    ).select_related('lesson', 'lesson__course').order_by('-mark_time')
    
    for row_idx, attendance in enumerate(attendance_records, 2):
        attendance_data = [
            attendance.mark_time.strftime('%d.%m.%Y') if attendance.mark_time else '',
            attendance.lesson.course.title if attendance.lesson and attendance.lesson.course else '',
            attendance.get_attendance_status_display(),
            getattr(attendance.lesson, 'lesson_type', '') if attendance.lesson else '',
            attendance.comment or '',
        ]
        
        for col_idx, value in enumerate(attendance_data, 1):
            cell = ws_attendance.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
    
    # Автоширина
    for column in ws_attendance.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_attendance.column_dimensions[column_letter].width = adjusted_width
    
    # 4. Лист с финансовой информацией
    ws_finance = wb.create_sheet('Финансы', 3)
    
    # Заголовки
    finance_headers = ['Дата', 'Курс', 'Тип операции', 'Сумма', 'Способ оплаты', 'Примечание']
    for col_idx, header in enumerate(finance_headers, 1):
        cell = ws_finance.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment
    
    # Оплаты
    payments = Payment.objects.filter(student=student).select_related('course').order_by('-paid_at')
    for row_idx, payment in enumerate(payments, 2):
        finance_data = [
            payment.paid_at.strftime('%d.%m.%Y') if payment.paid_at else '',
            payment.course.title if payment.course else '',
            'Оплата',
            payment.amount,
            payment.get_payment_method_display(),
            payment.comment or '',
        ]
        
        for col_idx, value in enumerate(finance_data, 1):
            cell = ws_finance.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
    
    # Долги
    debts = Debt.objects.filter(student=student).select_related('course')
    current_row = len(payments) + 3
    
    for row_idx, debt in enumerate(debts, current_row):
        debt_data = [
            debt.created_at.strftime('%d.%m.%Y') if debt.created_at else '',
            debt.course.title if debt.course else '',
            'Долг',
            debt.total_amount,
            '',
            f'Статус: {debt.get_status_display()}',
        ]
        
        for col_idx, value in enumerate(debt_data, 1):
            cell = ws_finance.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
    
    # Автоширина
    for column in ws_finance.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_finance.column_dimensions[column_letter].width = adjusted_width
    
    # 5. Лист с заметками менторов
    ws_notes = wb.create_sheet('Заметки менторов', 4)
    
    # Заголовки
    notes_headers = ['Дата', 'Ментор', 'Текст заметки']
    for col_idx, header in enumerate(notes_headers, 1):
        cell = ws_notes.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment
    
    # Заметки
    notes = StudentNote.objects.filter(student=student).select_related('mentor').order_by('-created_at')
    for row_idx, note in enumerate(notes, 2):
        note_data = [
            note.created_at.strftime('%d.%m.%Y %H:%M') if note.created_at else '',
            f'{note.mentor.first_name} {note.mentor.last_name}' if note.mentor else '',
            note.text or '',
        ]
        
        for col_idx, value in enumerate(note_data, 1):
            cell = ws_notes.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 3:  # Текст заметки
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Автоширина
    ws_notes.column_dimensions['A'].width = 20
    ws_notes.column_dimensions['B'].width = 25
    ws_notes.column_dimensions['C'].width = 80
    
    # 6. Лист с активностью и логами
    ws_activity = wb.create_sheet('Активность', 5)
    
    # Заголовки
    activity_headers = ['Параметр', 'Значение']
    for col_idx, header in enumerate(activity_headers, 1):
        cell = ws_activity.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment
    
    # Статистика активности
    total_attendance = attendance_records.count()
    present_attendance = attendance_records.filter(attendance_status='present').count()
    attendance_percentage = (present_attendance / total_attendance * 100) if total_attendance > 0 else 0
    
    activity_data = [
        ['Общее количество посещений', total_attendance],
        ['Присутствовал', present_attendance],
        ['Процент посещения', f'{attendance_percentage:.1f}%'],
        ['Количество курсов', enrollments.count()],
        ['Активные курсы', enrollments.filter(status='active').count()],
        ['Завершенные курсы', enrollments.filter(status='completed').count()],
        ['Общее количество оплат', payments.count()],
        ['Сумма всех оплат', payments.aggregate(total=Sum('amount'))['total'] or 0],
        ['Текущие долги', debts.filter(status='active').aggregate(total=Sum('total_amount'))['total'] or 0],
        ['Количество заметок', notes.count()],
        ['Последняя активность', student.user_account.last_login.strftime('%d.%m.%Y %H:%M') if student.user_account and student.user_account.last_login else ''],
    ]
    
    for row_idx, (param, value) in enumerate(activity_data, 2):
        ws_activity.cell(row=row_idx, column=1, value=param).border = border
        ws_activity.cell(row=row_idx, column=2, value=value).border = border
    
    # Автоширина
    ws_activity.column_dimensions['A'].width = 30
    ws_activity.column_dimensions['B'].width = 25
    
    # Создаём HTTP ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'detailed_report_{student.full_name.replace(" ", "_")}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


def student_info_page(request, student_id):
    """Страница с детальной информацией о студенте для родителей"""
    from django.db.models import Sum
    
    student = get_object_or_404(Student, pk=student_id)
    
    # Получаем заметки менторов
    notes = StudentNote.objects.filter(student=student).select_related('mentor').order_by('-created_at')
    
    # Получаем информацию о курсах, оплатах и оценках
    enrollments = student.course_enrollments.select_related('course').prefetch_related('course__mentor').all()
    
    # Добавляем информацию об оплатах для каждого курса
    for enrollment in enrollments:
        # Получаем оплаты для этого курса через студента
        total_paid = student.payments.filter(course=enrollment.course).aggregate(total=Sum('amount'))['total'] or 0
        enrollment.total_paid = total_paid
        enrollment.course_payments = student.payments.filter(course=enrollment.course)
    
    context = {
        'student': student,
        'notes': notes,
        'enrollments': enrollments,
    }
    
    return render(request, 'students/student_info_page.html', context)


@login_required
@admin_required
def add_admin_note(request, pk):
    """Добавление заметки администратора о студенте"""
    from .models import AdminStudentNote
    
    student = get_object_or_404(Student, pk=pk)
    
    if request.method == 'POST':
        note_text = request.POST.get('note_text')
        if note_text:
            # Создаем заметку администратора
            AdminStudentNote.objects.create(
                student=student,
                admin=request.user,
                text=note_text
            )
            messages.success(request, 'Заметка администратора добавлена')
        else:
            messages.error(request, 'Текст заметки не может быть пустым')
    
    # Перенаправляем обратно на страницу студента
    return redirect('admin_students:detail', pk=pk)
