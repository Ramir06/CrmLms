from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import logout
from django.contrib import messages as django_messages
from django.http import HttpResponseRedirect, Http404, HttpResponse
from django.urls import reverse_lazy
from django.utils import timezone
from django.db.models import Sum, Count, Avg, Q
from apps.courses.models import Course, CourseStudent
from apps.lectures.models import Section, Material
from apps.quizzes.models import Quiz, QuizAttempt
from apps.attendance.models import AttendanceRecord
from apps.lessons.models import Lesson
from apps.assignments.models import Assignment, AssignmentSubmission
from apps.notifications.models import Notification
from apps.news.models import News
from apps.news.forms import NewsForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import json
from io import BytesIO


def get_student_profile(user):
    """Get the Student model instance linked to this user."""
    if not user.is_student:
        return None
    return getattr(user, 'student_profile', None)


@login_required
def student_dashboard(request):
    """Student dashboard with their courses, recent activity, and stats."""
    if not request.user.is_student:
        return HttpResponseRedirect(reverse_lazy('dashboard:index'))

    student = get_student_profile(request.user)
    if not student:
        raise Http404('Student profile not found')

    # Get student's organization
    student_organization = getattr(student, 'organization', None)

    enrollments = CourseStudent.objects.filter(
        student=student,
        status='active'
    ).select_related('course', 'course__mentor').order_by('-joined_at')

    recent_quizzes = QuizAttempt.objects.filter(
        course_student__student=student
    ).select_related('quiz', 'course_student__course').order_by('-submitted_at')[:5]

    # Recent assignments (pending, across all courses)
    from apps.assignments.models import Assignment, AssignmentSubmission, AssignmentGrade
    active_course_ids = enrollments.values_list('course_id', flat=True)
    recent_assignments = Assignment.objects.filter(
        course_id__in=active_course_ids, is_visible=True
    ).select_related('course').order_by('-due_date', '-id')[:5]

    # Check which have submissions
    my_submissions = AssignmentSubmission.objects.filter(
        assignment__in=recent_assignments, student=student
    ).values_list('assignment_id', flat=True)
    submitted_ids = set(my_submissions)

    assignment_list = []
    for a in recent_assignments:
        assignment_list.append({
            'assignment': a,
            'submitted': a.pk in submitted_ids,
        })

    # Recent grades
    recent_grades = AssignmentGrade.objects.filter(
        submission__student=student,
        submission__assignment__course_id__in=active_course_ids,
    ).select_related(
        'submission__assignment', 'submission__assignment__course'
    ).order_by('-checked_at')[:5]

    # Notifications
    recent_notifications = Notification.objects.filter(
        recipient=request.user
    ).order_by('-created_at')[:3]
    
    unread_count = Notification.objects.filter(
        recipient=request.user,
        is_read=False
    ).count()

    context = {
        'enrollments': enrollments,
        'recent_quizzes': recent_quizzes,
        'recent_assignments': assignment_list,
        'recent_grades': recent_grades,
        'student': student,
        'student_organization': student_organization,
        'page_title': 'Мой кабинет',
        'recent_notifications': recent_notifications,
        'unread_count': unread_count,
    }
    return render(request, 'students/dashboard.html', context)


@login_required
def course_detail(request, course_id):
    """Student view of a single course with materials and progress."""
    if not request.user.is_student:
        return HttpResponseRedirect(reverse_lazy('dashboard:index'))

    student = get_student_profile(request.user)
    if not student:
        raise Http404('Student profile not found')

    enrollment = get_object_or_404(
        CourseStudent,
        student=student,
        course_id=course_id,
        status='active'
    )
    course = enrollment.course
    # Get sections with lectures and materials
    from apps.lectures.models import Lecture
    from apps.assignments.models import Assignment
    sections = Section.objects.filter(
        course=course,
        is_visible=True
    ).prefetch_related(
        'materials',
        'lectures__materials',
        'assignments',  # Добавляем загрузку заданий для разделов
        'quizzes'
    ).order_by('order')
    
    # Create a list of sections with filtered assignments
    sections_with_filtered_assignments = []
    for section in sections:
        # Get only visible assignments for this section
        visible_assignments = section.assignments.filter(is_visible=True)
        
        # Create a simple object to hold section data with filtered assignments
        section_data = {
            'pk': section.pk,
            'title': section.title,
            'order': section.order,
            'is_visible': section.is_visible,
            'materials': section.materials.all(),
            'lectures': section.lectures.all(),
            'assignments': visible_assignments,  # Already filtered
            'quizzes': section.quizzes.all(),
        }
        sections_with_filtered_assignments.append(section_data)

    # Get quiz attempts for this course
    attempts = QuizAttempt.objects.filter(
        course_student=enrollment
    ).select_related('quiz').order_by('-submitted_at')

    context = {
        'course': course,
        'enrollment': enrollment,
        'sections': sections_with_filtered_assignments,  # Use filtered sections
        'attempts': attempts,
        'page_title': course.title,
    }
    return render(request, 'students/course_lectures.html', context)


@login_required
def material_detail(request, course_id, material_id):
    """Student view of a specific material."""
    if not request.user.is_student:
        return HttpResponseRedirect(reverse_lazy('dashboard:index'))

    student = get_student_profile(request.user)
    if not student:
        raise Http404('Student profile not found')

    enrollment = get_object_or_404(
        CourseStudent,
        student=student,
        course_id=course_id,
        status='active'
    )
    material = get_object_or_404(
        Material,
        pk=material_id,
        section__course_id=course_id,
        is_visible=True
    )
    course = material.section.course

    # Prev/next navigation within the same section
    section_materials = Material.objects.filter(
        section=material.section, is_visible=True
    ).order_by('order', 'id')
    prev_material = section_materials.filter(order__lt=material.order).last()
    next_material = section_materials.filter(order__gt=material.order).first()

    context = {
        'course': course,
        'enrollment': enrollment,
        'material': material,
        'prev_material': prev_material,
        'next_material': next_material,
        'page_title': material.title,
    }
    return render(request, 'students/material_detail.html', context)


@login_required
def material_detail_ajax(request, course_id, material_id):
    """AJAX: render material detail for right panel (student version)."""
    if not request.user.is_student:
        return HttpResponseRedirect(reverse_lazy('dashboard:index'))

    student = get_student_profile(request.user)
    if not student:
        raise Http404('Student profile not found')

    enrollment = get_object_or_404(
        CourseStudent,
        student=student,
        course_id=course_id,
        status='active'
    )
    material = get_object_or_404(
        Material,
        pk=material_id,
        section__course_id=course_id,
        is_visible=True
    )
    course = material.section.course

    context = {
        'course': course,
        'material': material,
        'section': material.section,
    }
    
    return render(request, 'students/material_detail_ajax.html', context)


def student_logout(request):
    """Logout for students."""
    logout(request)
    return HttpResponseRedirect(reverse_lazy('accounts:login'))


@login_required
def student_attendance(request, course_id):
    """Student view of their attendance for a specific course."""
    if not request.user.is_student:
        return HttpResponseRedirect(reverse_lazy('dashboard:index'))

    student = get_student_profile(request.user)
    if not student:
        raise Http404('Student profile not found')

    enrollment = get_object_or_404(
        CourseStudent,
        student=student,
        course_id=course_id,
        status='active'
    )
    course = enrollment.course

    # Get attendance records
    attendance_records = AttendanceRecord.objects.filter(
        lesson__course=course,
        student=student
    ).select_related('lesson').order_by('lesson__lesson_date', 'lesson__start_time')

    context = {
        'course': course,
        'enrollment': enrollment,
        'attendance_records': attendance_records,
        'page_title': 'Посещаемость',
    }
    return render(request, 'students/course_attendance.html', context)


@login_required
def student_grades(request, course_id):
    """Student view of their grades for a specific course."""
    if not request.user.is_student:
        return HttpResponseRedirect(reverse_lazy('dashboard:index'))

    student = get_student_profile(request.user)
    if not student:
        raise Http404('Student profile not found')

    enrollment = get_object_or_404(
        CourseStudent,
        student=student,
        course_id=course_id,
        status='active'
    )
    course = enrollment.course

    # Get grades for this course
    from apps.assignments.models import AssignmentGrade
    grades = AssignmentGrade.objects.filter(
        submission__student=student,
        submission__assignment__course=course
    ).select_related('submission__assignment').order_by('-checked_at')

    # Calculate statistics
    grade_scores = [g.score for g in grades if g.score is not None]
    average_grade = round(sum(grade_scores) / len(grade_scores), 1) if grade_scores else None
    max_grade = max(grade_scores) if grade_scores else None
    passed_assignments = [g for g in grades if g.score is not None and g.score >= 80]

    context = {
        'course': course,
        'enrollment': enrollment,
        'grades': grades,
        'average_grade': average_grade,
        'max_grade': max_grade,
        'passed_assignments': passed_assignments,
        'page_title': 'Оценки',
    }
    return render(request, 'students/course_grades.html', context)


def _get_student_or_redirect(request):
    """Helper: returns (student, None) or (None, redirect_response)."""
    if not request.user.is_student:
        return None, HttpResponseRedirect(reverse_lazy('dashboard:index'))
    student = get_student_profile(request.user)
    if not student:
        raise Http404('Student profile not found')
    return student, None


def _get_enrollment(student, course_id):
    """Helper: returns enrollment for student + course."""
    return get_object_or_404(
        CourseStudent, student=student, course_id=course_id, status='active'
    )


# ─── Schedule ───────────────────────────────────────────────────
@login_required
def student_schedule(request):
    """Student schedule: diary format with day navigation."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    enrollments = CourseStudent.objects.filter(
        student=student, status='active'
    ).select_related('course')
    course_ids = [e.course_id for e in enrollments]

    # Получаем выбранную дату или сегодня
    date_str = request.GET.get('date')
    if date_str:
        try:
            selected_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.localdate()
    else:
        selected_date = timezone.localdate()
    
    # Получаем начало недели (понедельник)
    start_of_week = selected_date - timezone.timedelta(days=selected_date.weekday())
    
    # Создаем информацию о днях недели
    week_days = []
    day_names = ['ПН', 'ВТ', 'СР', 'ЧТ', 'ПТ', 'СБ', 'ВС']
    month_names = ['', 'ЯНВ', 'ФЕВ', 'МАР', 'АПР', 'МАЙ', 'ИЮН', 'ИЮЛ', 'АВГ', 'СЕН', 'ОКТ', 'НОЯ', 'ДЕК']
    
    for i in range(7):
        current_day = start_of_week + timezone.timedelta(days=i)
        day_lessons = Lesson.objects.filter(
            course_id__in=course_ids,
            lesson_date=current_day,
            status__in=['scheduled', 'completed'],
        ).select_related('course', 'temporary_mentor').order_by('start_time')
        
        week_days.append({
            'date': current_day,
            'day_name': day_names[i],
            'is_today': current_day == timezone.localdate(),
            'lessons_count': day_lessons.count(),
        })
    
    # Получаем уроки для выбранного дня
    day_lessons = Lesson.objects.filter(
        course_id__in=course_ids,
        lesson_date=selected_date,
        status__in=['scheduled', 'completed'],
    ).select_related('course', 'temporary_mentor').order_by('start_time')
    
    # Добавляем цвета и дополнительную информацию
    colors = ['#9EABC5', '#E9E6E0', '#A7D6AA', '#F3F7C1', '#CA9CAC', '#B5D8EB', '#F5B7B1']
    for i, lesson in enumerate(day_lessons):
        lesson.color = colors[i % len(colors)]
        lesson.homework = getattr(lesson, 'homework', 'Не задано')

    context = {
        'week_days': week_days,
        'day_lessons': day_lessons,
        'selected_date': selected_date,
        'page_title': 'Расписание',
    }
    return render(request, 'students/schedule.html', context)


# ─── Profile ─────────────────────────────────────────────────────
@login_required
def student_profile(request):
    """Student profile page with settings."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    # Получаем статистику студента
    enrolled_courses = CourseStudent.objects.filter(student=student, status='active')
    completed_assignments = AssignmentSubmission.objects.filter(
        assignment__course__in=enrolled_courses.values('course'),
        student=student,
        status='submitted'
    ).count()
    
    # Расчет посещаемости
    total_lessons = Lesson.objects.filter(
        course__in=enrolled_courses.values('course'),
        status='completed'
    ).count()
    
    attended_lessons = AttendanceRecord.objects.filter(
        lesson__course__in=enrolled_courses.values('course'),
        student=student,
        attendance_status='present'
    ).count()
    
    attendance_rate = int((attended_lessons / total_lessons * 100) if total_lessons > 0 else 0)

    context = {
        'student': student,
        'enrolled_courses': enrolled_courses,
        'completed_assignments': completed_assignments,
        'attendance_rate': attendance_rate,
        'page_title': 'Профиль',
    }
    return render(request, 'students/profile.html', context)


@login_required
def profile_update(request):
    """Update student profile information."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    if request.method == 'POST':
        try:
            # Обновляем информацию пользователя
            if student.user_account:
                user = student.user_account
                user.first_name = request.POST.get('first_name', user.first_name)
                user.last_name = request.POST.get('last_name', user.last_name)
                user.save()
            
            # Обновляем информацию студента
            student.phone = request.POST.get('phone', student.phone)
            student.bio = request.POST.get('bio', student.bio)
            
            birth_date = request.POST.get('birth_date')
            if birth_date:
                student.birth_date = datetime.datetime.strptime(birth_date, '%Y-%m-%d').date()
            
            student.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def settings_update(request):
    """Update student notification settings."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    if request.method == 'POST':
        try:
            # Обновляем настройки уведомлений
            email_notifications = {
                'assignments': request.POST.get('email_assignments') == 'on',
                'grades': request.POST.get('email_grades') == 'on',
                'schedule': request.POST.get('email_schedule') == 'on',
                'news': request.POST.get('email_news') == 'on',
            }
            
            student.email_notifications = email_notifications
            student.language = request.POST.get('language', 'ru')
            student.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def password_change(request):
    """Change student password."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    if request.method == 'POST':
        try:
            current_password = request.POST.get('current_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            # Проверяем текущий пароль
            if not student.user_account or not student.user_account.check_password(current_password):
                return JsonResponse({'success': False, 'error': 'Неверный текущий пароль'})
            
            # Проверяем совпадение новых паролей
            if new_password != confirm_password:
                return JsonResponse({'success': False, 'error': 'Пароли не совпадают'})
            
            # Проверяем длину пароля
            if len(new_password) < 8:
                return JsonResponse({'success': False, 'error': 'Пароль должен содержать минимум 8 символов'})
            
            # Обновляем пароль
            student.user_account.set_password(new_password)
            student.user_account.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def avatar_upload(request):
    """Upload student avatar."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    if request.method == 'POST' and request.FILES.get('avatar'):
        try:
            avatar = request.FILES['avatar']
            
            # Проверяем размер файла (макс 5MB)
            if avatar.size > 5 * 1024 * 1024:
                return JsonResponse({'success': False, 'error': 'Размер файла не должен превышать 5MB'})
            
            # Проверяем тип файла
            allowed_types = ['image/jpeg', 'image/png', 'image/gif']
            if avatar.content_type not in allowed_types:
                return JsonResponse({'success': False, 'error': 'Разрешены только JPG, PNG и GIF файлы'})
            
            # Удаляем старый аватар если есть
            if student.profile_image:
                student.profile_image.delete()
            
            # Сохраняем новый аватар
            student.profile_image = avatar
            student.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def two_factor_toggle(request):
    """Toggle two-factor authentication."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            enabled = data.get('enabled', False)
            
            student.two_factor_enabled = enabled
            student.save()
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def get_lesson_number(start_time):
    """Определяет номер урока по времени начала"""
    lesson_times = {
        1: (8, 30),   # 8:30
        2: (10, 15),  # 10:15
        3: (12, 0),   # 12:00
        4: (13, 45),  # 13:45
        5: (15, 30),  # 15:30
        6: (17, 15),  # 17:15
        7: (19, 0),   # 19:00
        8: (20, 45),  # 20:45
    }
    
    for num, (hour, minute) in lesson_times.items():
        if start_time.hour == hour and start_time.minute == minute:
            return num
    return None


# ─── Lesson detail ──────────────────────────────────────────────
@login_required
def student_lesson_detail(request, course_id, lesson_id):
    """Student view of a single lesson: topic, description, meet link, attendance."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    enrollment = _get_enrollment(student, course_id)
    lesson = get_object_or_404(Lesson, pk=lesson_id, course_id=course_id)

    att_record = AttendanceRecord.objects.filter(
        lesson=lesson, student=student
    ).first()

    context = {
        'course': enrollment.course,
        'enrollment': enrollment,
        'lesson': lesson,
        'att_record': att_record,
        'page_title': lesson.title or f'Занятие {lesson.lesson_date}',
    }
    return render(request, 'students/lesson_detail.html', context)


@login_required
def student_lesson_detail_ajax(request, course_id, lesson_id):
    """AJAX version of student lesson detail."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    enrollment = _get_enrollment(student, course_id)
    lesson = get_object_or_404(Lesson, pk=lesson_id, course_id=course_id)

    att_record = AttendanceRecord.objects.filter(
        lesson=lesson, student=student
    ).first()

    context = {
        'course': enrollment.course,
        'enrollment': enrollment,
        'lesson': lesson,
        'att_record': att_record,
        'page_title': lesson.title or f'Занятие {lesson.lesson_date}',
    }
    
    # Возвращаем только контент урока для AJAX
    return render(request, 'students/partials/lesson_content.html', context)


# ─── Assignments list ───────────────────────────────────────────
@login_required
def student_assignments(request, course_id):
    """Student list of all assignments for a course with submission status."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    enrollment = _get_enrollment(student, course_id)
    course = enrollment.course

    assignments = Assignment.objects.filter(
        course=course, is_visible=True
    ).order_by('order', 'id')
    
    submissions = AssignmentSubmission.objects.filter(
        assignment__in=assignments, student=student
    ).select_related('grade')
    sub_map = {s.assignment_id: s for s in submissions}

    today = timezone.localdate()
    assignments_data = []
    for a in assignments:
        sub = sub_map.get(a.pk)
        grade_obj = None
        if sub:
            try:
                grade_obj = sub.grade
            except Exception:
                pass
        assignments_data.append({
            'assignment': a,
            'submission': sub,
            'grade': grade_obj,
            'status': sub.status if sub else 'not_submitted',
            'is_overdue': a.due_date and a.due_date < today and (not sub or sub.status == 'not_submitted'),
        })

    context = {
        'course': course,
        'enrollment': enrollment,
        'assignments_data': assignments_data,
        'page_title': 'Задания',
    }
    return render(request, 'students/assignments.html', context)


# ─── Assignment submit ──────────────────────────────────────────
@login_required
def student_assignment_submit(request, course_id, assignment_id):
    """Student submits (or re-submits) an assignment."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    enrollment = _get_enrollment(student, course_id)
    course = enrollment.course
    assignment = get_object_or_404(
        Assignment, pk=assignment_id, course=course, is_visible=True
    )

    submission, created = AssignmentSubmission.objects.get_or_create(
        assignment=assignment, student=student,
        defaults={'status': 'not_submitted'}
    )

    if request.method == 'POST':
        answer_text = request.POST.get('answer_text', '').strip()
        file = request.FILES.get('file')
        if answer_text or file:
            submission.answer_text = answer_text
            if file:
                submission.file = file
            submission.status = 'submitted'
            submission.submitted_at = timezone.now()
            submission.save()
            django_messages.success(request, 'Задание отправлено!')
            return redirect('students:assignments', course_id=course_id)
        else:
            django_messages.error(request, 'Заполните ответ или прикрепите файл.')

    context = {
        'course': course,
        'enrollment': enrollment,
        'assignment': assignment,
        'submission': submission,
        'page_title': assignment.title,
    }
    return render(request, 'students/assignment_submit.html', context)


@login_required
def student_assignment_detail_ajax(request, course_id, assignment_id):
    """AJAX version of student assignment detail."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    enrollment = _get_enrollment(student, course_id)
    course = enrollment.course
    assignment = get_object_or_404(
        Assignment, pk=assignment_id, course=course, is_visible=True
    )

    submission, created = AssignmentSubmission.objects.get_or_create(
        assignment=assignment, student=student,
        defaults={'status': 'not_submitted'}
    )

    context = {
        'course': enrollment.course,
        'enrollment': enrollment,
        'assignment': assignment,
        'submission': submission,
        'page_title': assignment.title,
    }
    
    # Возвращаем только контент задания для AJAX
    return render(request, 'students/partials/assignment_content.html', context)


# ─── Quizzes list ───────────────────────────────────────────────
@login_required
def student_quizzes(request, course_id):
    """Student list of all quizzes for a course."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    enrollment = _get_enrollment(student, course_id)
    course = enrollment.course

    quizzes = Quiz.objects.filter(course=course, is_active=True).order_by('-created_at')

    attempts = QuizAttempt.objects.filter(
        course_student=enrollment
    ).select_related('quiz')
    attempt_map = {}
    for att in attempts:
        attempt_map.setdefault(att.quiz_id, []).append(att)

    quizzes_data = []
    for q in quizzes:
        q_attempts = attempt_map.get(q.pk, [])
        best = max(q_attempts, key=lambda a: a.percentage, default=None) if q_attempts else None
        quizzes_data.append({
            'quiz': q,
            'attempts_count': len(q_attempts),
            'best': best,
        })

    context = {
        'course': course,
        'enrollment': enrollment,
        'quizzes_data': quizzes_data,
        'page_title': 'Тесты',
    }
    return render(request, 'students/quizzes.html', context)


# ─── Quiz start ─────────────────────────────────────────────────
@login_required
def student_quiz_start(request, course_id, quiz_id):
    """Create a new quiz attempt and redirect to quiz-taking page."""
    student, redir = _get_student_or_redirect(request)
    if redir:
        return redir

    enrollment = _get_enrollment(student, course_id)
    quiz = get_object_or_404(Quiz, pk=quiz_id, course_id=course_id, is_active=True)

    # Create new attempt
    attempt = QuizAttempt.objects.create(
        quiz=quiz,
        course_student=enrollment,
        max_score=quiz.total_points(),
    )
    return redirect('quiz_take', token=attempt.token)


@login_required
def student_notifications(request):
    """Student notifications page."""
    if not request.user.is_student:
        return HttpResponseRedirect(reverse_lazy('dashboard:index'))

    student = get_student_profile(request.user)
    if not student:
        raise Http404('Student profile not found')

    # Get student's organization
    student_organization = getattr(student, 'organization', None)

    notifications = Notification.objects.filter(
        recipient=request.user
    ).order_by('-created_at')

    # Mark all as read when viewing the page
    notifications.filter(is_read=False).update(is_read=True)

    context = {
        'notifications': notifications,
        'student_organization': student_organization,
        'page_title': 'Уведомления',
    }
    return render(request, 'students/notifications.html', context)


@login_required
def student_payments(request):
    """Student payments page."""
    if not request.user.is_student:
        return HttpResponseRedirect(reverse_lazy('dashboard:index'))

    student = get_student_profile(request.user)
    if not student:
        raise Http404('Student profile not found')

    # Get student's organization
    student_organization = getattr(student, 'organization', None)

    enrollments = CourseStudent.objects.filter(
        student=student,
        status='active'
    ).select_related('course', 'course__mentor').order_by('-joined_at')

    # Add payment progress calculation
    enrollment_data = []
    for enrollment in enrollments:
        # Get course duration in months (default to 4 if not set)
        course_months = getattr(enrollment.course, 'duration_months', 4) or 4
        
        # Use the existing paid_months property from CourseStudent
        paid_months = enrollment.paid_months
        
        # Calculate payment percentage
        payment_percentage = (paid_months / course_months * 100) if course_months > 0 else 0
        
        # Get total payments amount
        from apps.payments.models import Payment
        payments = Payment.objects.filter(
            student=student,
            course=enrollment.course
        ).aggregate(
            total_amount=Sum('amount')
        )
        total_paid = payments['total_amount'] or 0
        
        # Create month list for template
        month_list = list(range(1, course_months + 1))
        
        enrollment_data.append({
            'enrollment': enrollment,
            'payment_percentage': int(payment_percentage),
            'paid_months': paid_months,
            'course_months': course_months,
            'month_list': month_list,
            'paid_amount': total_paid,
            'total_amount': None,  # Will be calculated based on course price if needed
        })

    context = {
        'enrollment_data': enrollment_data,
        'student_organization': student_organization,
        'page_title': 'Оплата за курс',
    }
    return render(request, 'students/payments.html', context)


@login_required
def export_excel_report(request):
    """Export comprehensive Excel report with student statistics."""
    if not request.user.is_student:
        return HttpResponseRedirect(reverse_lazy('dashboard:index'))

    student = get_student_profile(request.user)
    if not student:
        raise Http404('Student profile not found')

    # Get student's organization
    student_organization = getattr(student, 'organization', None)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Report number (timestamp based)
    report_number = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Sheet 1: По Филлиалам
    ws1 = wb.create_sheet("По Филлиалам")
    
    # Headers
    headers1 = ['Филлиал', 'Учеников', 'Менторов', 'Курсов']
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    
    # Data for branches
    if student_organization:
        branch_data = [
            [student_organization.name or 'ЦУМ', 
             CourseStudent.objects.filter(student__organization=student_organization).count(),
             Course.objects.filter(mentor__organization=student_organization).count(),
             Course.objects.filter(organization=student_organization).count()]
        ]
        
        for row_idx, row_data in enumerate(branch_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin")
                )
    
    # Auto-adjust column widths
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 2: По лидам
    ws2 = wb.create_sheet("По лидам")
    
    # Headers
    headers2 = ['Филлиал', 'Количество', 'Конверсия %', 'Рейтинг']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    
    # Data for leads
    if student_organization:
        total_leads = 464  # Example data - you can calculate this from your leads model
        conversion_rate = 2  # Example conversion rate
        
        leads_data = [
            [student_organization.name or 'ЦУМ', total_leads, conversion_rate, '1-е место']
        ]
        
        for row_idx, row_data in enumerate(leads_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin")
                )
    
    # Auto-adjust column widths
    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 3: По отзывам студентов
    ws3 = wb.create_sheet("По отзывам студентов")
    
    # Headers
    headers3 = ['Работа ментора', 'Самооценка']
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    
    # Data for feedback
    # Calculate average mentor rating and self-assessment
    enrollments = CourseStudent.objects.filter(student=student)
    avg_mentor_rating = 57  # Example data - you can calculate from feedback model
    self_assessment = 90   # Example data - you can calculate from self-assessment model
    
    feedback_data = [
        [f"{avg_mentor_rating}/100", f"{self_assessment}/100"]
    ]
    
    for row_idx, row_data in enumerate(feedback_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
    
    # Auto-adjust column widths
    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 20
    
    # Save to memory
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="report_{report_number}.xlsx"'
    
    return response


@login_required
def news_list_student(request):
    """Список новостей для студентов"""
    if not request.user.is_student:
        return HttpResponseRedirect(reverse_lazy('dashboard:index'))
    
    student = get_student_profile(request.user)
    if not student:
        raise Http404('Student profile not found')
    
    # Фильтруем по организации студента
    organization = getattr(student, 'organization', None)
    
    # Получаем новости для студентов
    news_query = News.objects.filter(is_published=True)
    
    if organization:
        news_query = news_query.filter(
            Q(organization=organization) | Q(organization__isnull=True)
        )
    
    # Фильтруем по аудитории
    news_query = news_query.filter(
        Q(audience__in=['all', 'students'])
    )
    
    news = news_query.select_related('created_by').order_by('-published_at')
    
    return render(request, 'students/news/list.html', {
        'news': news, 
        'page_title': 'Объявления'
    })


@login_required
def news_detail_student(request, pk):
    """Детальная новость для студентов"""
    if not request.user.is_student:
        return HttpResponseRedirect(reverse_lazy('dashboard:index'))
    
    student = get_student_profile(request.user)
    if not student:
        raise Http404('Student profile not found')
    
    # Фильтруем по организации студента
    organization = getattr(student, 'organization', None)
    
    # Получаем новость
    news_query = News.objects.filter(is_published=True, pk=pk)
    
    if organization:
        news_query = news_query.filter(
            Q(organization=organization) | Q(organization__isnull=True)
        )
    
    news = get_object_or_404(news_query, audience__in=['all', 'students'])
    
    return render(request, 'students/news/detail.html', {
        'news': news, 
        'page_title': news.title
    })
